import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== 色定義 =====
COLOR_HEADER_HOSPITAL = "1F4E79"   # 濃い青（病院名ヘッダー）
COLOR_HEADER_CATEGORY = "2E75B6"   # 中青（カテゴリヘッダー）
COLOR_HEADER_COL = "D6E4F0"        # 薄青（列ヘッダー）
COLOR_UNKNOWN = "FFF2CC"           # 黄色（未確認）
COLOR_KNOWN = "E2EFDA"             # 薄緑（確認済み）

def thin_border():
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)

def header_font(size=11, bold=True, color="FFFFFF"):
    return Font(name="メイリオ", size=size, bold=bold, color=color)

def body_font(size=10, bold=False):
    return Font(name="メイリオ", size=size, bold=bold)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center_align(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

# ========== シート作成 ==========
ws = wb.active
ws.title = "放射線機器一覧"

# 列幅設定
col_widths = {
    "A": 18,  # 病院名
    "B": 20,  # カテゴリ
    "C": 10,  # 室番号
    "D": 22,  # メーカー名
    "E": 28,  # 型名
    "F": 14,  # 導入年月
    "G": 30,  # 備考
}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# 行の高さ
ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 18

# ========== タイトル行 ==========
ws.merge_cells("A1:G1")
title_cell = ws["A1"]
title_cell.value = "放射線機器一覧（一般撮影・X線TV・血管撮影）"
title_cell.font = Font(name="メイリオ", size=14, bold=True, color="FFFFFF")
title_cell.fill = fill(COLOR_HEADER_HOSPITAL)
title_cell.alignment = center_align()
title_cell.border = thin_border()

# ========== 列ヘッダー ==========
headers = ["病院名", "検査種別", "室・号機", "メーカー名", "型名", "導入年月", "備考"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_idx, value=header)
    cell.font = header_font(size=10, color="1F4E79")
    cell.fill = fill(COLOR_HEADER_COL)
    cell.alignment = center_align(wrap=True)
    cell.border = thin_border()

# ========== データ定義 ==========
# 凡例: (病院名, カテゴリ, 室番号, メーカー, 型名, 導入年月, 備考, 確認済みフラグ)
# 確認済みフラグ: True=確認済み(緑), False=未確認(黄)
data = [
    # ─── 北海道医療センター ─────────────────────────────────────
    ("北海道医療センター", "一般撮影", "第1撮影室",
     "コニカミノルタ", "AeroDR（FPD）＋撮影台",
     "不明",
     "FPD：AeroDRシリーズ使用確認。具体的モデル・導入年月は要確認", False),

    ("北海道医療センター", "一般撮影", "第2撮影室",
     "コニカミノルタ", "AeroDR（FPD）＋撮影台",
     "2022年9月",
     "2022年9月に更新（動態撮影対応）", True),

    ("北海道医療センター", "一般撮影", "第3撮影室",
     "コニカミノルタ", "AeroDR（FPD）＋撮影台",
     "2022年9月",
     "2022年9月に更新（動態撮影対応）", True),

    ("北海道医療センター", "一般撮影", "第4撮影室",
     "コニカミノルタ", "AeroDR（FPD）＋撮影台",
     "2022年2月",
     "2022年2月に更新（動態撮影対応）", True),

    ("北海道医療センター", "一般撮影", "第5撮影室",
     "コニカミノルタ", "AeroDR（FPD）＋撮影台",
     "2022年2月",
     "2022年2月に更新（動態撮影対応）", True),

    ("北海道医療センター", "透視撮影台（X線TV）", "第1透視室（汎用型）",
     "未確認", "未確認",
     "2022年9月",
     "2022年9月に汎用型へ更新。メーカー・型名は要確認", False),

    ("北海道医療センター", "透視撮影台（X線TV）", "第2透視室（Cアーム型）",
     "未確認", "未確認",
     "2022年9月",
     "2022年9月にCアーム型へ更新。メーカー・型名は要確認", False),

    ("北海道医療センター", "血管撮影（アンギオ）", "第1血管撮影室",
     "未確認", "未確認",
     "未確認",
     "公式サイト(403エラー)、WEB検索でも情報取得不可。要問合せ", False),

    # ─── 岸和田徳洲会病院 ──────────────────────────────────────
    ("岸和田徳洲会病院", "一般撮影", "撮影室",
     "未確認", "未確認",
     "未確認",
     "公式サイト(403エラー)、WEB検索でも情報取得不可。要問合せ", False),

    ("岸和田徳洲会病院", "透視撮影台（X線TV）", "第1透視室",
     "未確認", "未確認",
     "未確認",
     "X線TV装置は3台稼働確認済み。メーカー・型名・導入年月は要確認", False),

    ("岸和田徳洲会病院", "透視撮影台（X線TV）", "第2透視室",
     "未確認", "未確認",
     "未確認",
     "X線TV装置は3台稼働確認済み。メーカー・型名・導入年月は要確認", False),

    ("岸和田徳洲会病院", "透視撮影台（X線TV）", "第3透視室",
     "未確認", "未確認",
     "未確認",
     "X線TV装置は3台稼働確認済み。メーカー・型名・導入年月は要確認", False),

    ("岸和田徳洲会病院", "血管撮影（アンギオ）", "第1血管撮影室",
     "未確認", "未確認",
     "未確認",
     "血管造影装置は4台稼働確認済み。メーカー・型名・導入年月は要確認", False),

    ("岸和田徳洲会病院", "血管撮影（アンギオ）", "第2血管撮影室",
     "未確認", "未確認",
     "未確認",
     "血管造影装置は4台稼働確認済み。メーカー・型名・導入年月は要確認", False),

    ("岸和田徳洲会病院", "血管撮影（アンギオ）", "第3血管撮影室",
     "未確認", "未確認",
     "未確認",
     "血管造影装置は4台稼働確認済み。メーカー・型名・導入年月は要確認", False),

    ("岸和田徳洲会病院", "血管撮影（アンギオ）", "第4血管撮影室",
     "未確認", "未確認",
     "未確認",
     "血管造影装置は4台稼働確認済み。メーカー・型名・導入年月は要確認", False),
]

# ========== データ書き込み ==========
# 病院名を結合するための追跡
row = 3
hospital_start = {}
current_hospital = None

for i, (hospital, category, room, maker, model, date, note, confirmed) in enumerate(data):
    bg = COLOR_KNOWN if confirmed else COLOR_UNKNOWN

    vals = [hospital, category, room, maker, model, date, note]
    for col_idx, val in enumerate(vals, 1):
        cell = ws.cell(row=row + i, column=col_idx, value=val)
        cell.font = body_font()
        cell.border = thin_border()
        if col_idx == 1:
            cell.alignment = center_align(wrap=True)
            cell.font = body_font(bold=True)
        elif col_idx in (2, 3, 6):
            cell.alignment = center_align(wrap=True)
        else:
            cell.alignment = left_align(wrap=True)

        if col_idx in (4, 5, 6):
            cell.fill = fill(bg)

    ws.row_dimensions[row + i].height = 22

# ========== 病院名セルを縦結合 ==========
# 北海道医療センター: rows 3-10 (index 0-7)
ws.merge_cells(f"A3:A10")
for r in range(3, 11):
    ws[f"A{r}"].alignment = center_align(wrap=True)
    ws[f"A{r}"].font = body_font(bold=True)
    ws[f"A{r}"].fill = fill("DEEAF1")
    ws[f"A{r}"].border = thin_border()

# 岸和田徳洲会病院: rows 11-18 (index 8-15)
ws.merge_cells(f"A11:A18")
for r in range(11, 19):
    ws[f"A{r}"].alignment = center_align(wrap=True)
    ws[f"A{r}"].font = body_font(bold=True)
    ws[f"A{r}"].fill = fill("FBE5D6")
    ws[f"A{r}"].border = thin_border()

# ========== 注記シート ==========
ws2 = wb.create_sheet("注記・調査概要")
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 60

notes = [
    ("作成日", "2026年5月17日"),
    ("調査方法", "各病院公式ウェブサイト + WEB検索"),
    ("", ""),
    ("【北海道医療センター】", ""),
    ("公式URL", "https://hokkaido-mc.hosp.go.jp/each/radiology.html"),
    ("アクセス結果", "403 Forbidden（直接アクセス不可）"),
    ("確認できた情報",
     "・一般撮影：コニカミノルタ AeroDR（FPD）使用を確認\n"
     "・第2・3撮影室：2022年9月更新（動態撮影対応）\n"
     "・第4・5撮影室：2022年2月更新（動態撮影対応）\n"
     "・X線TV：2022年9月に2台更新（汎用型+Cアーム型）\n"
     "・血管撮影：詳細不明"),
    ("要確認事項",
     "一般撮影の具体的型名、X線TV/血管撮影のメーカー・型名・導入年月"),
    ("", ""),
    ("【岸和田徳洲会病院】", ""),
    ("公式URL", "https://kishiwada.tokushukai.or.jp/sp/section/radiolory/features.php"),
    ("アクセス結果", "403 Forbidden（直接アクセス不可）"),
    ("確認できた情報",
     "・血管造影装置（アンギオ）：4台稼働\n"
     "・X線TV装置：3台稼働\n"
     "・CT：3台、MRI：2台、PET-CT：1台"),
    ("要確認事項",
     "一般撮影・X線TV・血管撮影 全装置のメーカー・型名・導入年月"),
    ("", ""),
    ("推奨対応",
     "各病院の放射線部へ直接お問い合わせ、または病院ウェブサイトをブラウザで直接ご確認ください。\n"
     "※本ツールからは403エラーにより取得不可でした。"),
]

for r_idx, (key, val) in enumerate(notes, 1):
    ws2.row_dimensions[r_idx].height = 30
    cell_k = ws2.cell(row=r_idx, column=1, value=key)
    cell_v = ws2.cell(row=r_idx, column=2, value=val)
    cell_k.font = Font(name="メイリオ", size=10, bold=True)
    cell_k.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cell_v.font = Font(name="メイリオ", size=10)
    cell_v.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    if key.startswith("【"):
        cell_k.fill = fill("2E75B6")
        cell_k.font = Font(name="メイリオ", size=10, bold=True, color="FFFFFF")
        ws2.row_dimensions[r_idx].height = 20

# 保存
output_path = "/home/user/claude-code-on-the-web-html/放射線機器一覧.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")

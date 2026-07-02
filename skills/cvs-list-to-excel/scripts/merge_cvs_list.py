#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CVS納入リストPDF → Excel装置台帳 突合・上書きスクリプト

使い方:
    python3 merge_cvs_list.py --excel 台帳.xlsx --data transcription.json --out 出力.xlsx

transcription.json はPDFを精読して人手(モデル)が転記したデータ。形式:
{
  "prefecture": "沖縄県",          // Excelの「県」列の値と完全一致させる
  "year": 2026,                    // リストの年度
  "aliases": {                     // PDF病院名 → Excel病院名(正確な文字列) の対応表。
    "琉球大学病院": "琉球大学医学部附属病院"   // 自動マッチしない病院のみ書く
  },
  "hospitals": [
    {"name": "石垣島徳洲会病院",
     "devices": [
       {"maker": "フ", "system": "Azurion 3 M15", "plane": "S", "mark": "U"}
     ]}
  ]
}
maker: フ/キ/シ/島/G またはメーカー名。plane: "B"(バイ)/"S"(シングル)/""。
mark: "U"(Ⓤ更新)/"C"(Ⓒ既設)/"N"(Ⓝ新規)/""。

処理ルール(ユーザー指定の3ルール):
 1. システム名が同じ            → 行はそのまま
 2. システム名が異なる(装置更新) → メーカー/システム/納入年度(=year)を書換え。
                                   前装置に旧情報転記、行を赤文字、TYPE/主目的は空白+黄色
 3. ExcelにあるがPDFにない       → 備考に「(year)年リスト未掲載」を追記
PDFにあってExcelに無い装置・病院は行追加せず、質問列とレポートで顕在化する。
"""
import argparse
import difflib
import json
import sys
import unicodedata
from collections import Counter, defaultdict
from copy import copy

import openpyxl
from openpyxl.styles import Font, PatternFill

RED = "FFFF0000"
YELLOW = PatternFill(fill_type="solid", fgColor="FFFFFF00")

# メーカー略号 → 正規化キー
MAKER_KEY = {
    "フ": "PHILIPS", "ﾌｨﾘｯﾌﾟｽ": "PHILIPS", "フィリップス": "PHILIPS", "PHILIPS": "PHILIPS",
    "キ": "CANON", "ｷｬﾉﾝ": "CANON", "キヤノン": "CANON", "キャノン": "CANON", "CANON": "CANON",
    "シ": "SIEMENS", "ｼｰﾒﾝｽ": "SIEMENS", "シーメンス": "SIEMENS", "SIEMENS": "SIEMENS",
    "島": "SHIMADZU", "島津": "SHIMADZU", "SHIMADZU": "SHIMADZU",
    "G": "GE", "GE": "GE",
}


def norm(s):
    """全半角統一・空白除去・大文字化した比較用キー"""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    return "".join(s.split()).upper()


def maker_key(s):
    n = norm(s)
    return MAKER_KEY.get(str(s).strip(), MAKER_KEY.get(n, n))


def same_system(pdf_name, excel_name):
    """ルール1の「同じ」判定: 正規化後の一致・包含、シリーズ表記の前方一致"""
    p, e = norm(pdf_name), norm(excel_name)
    if not p or not e:
        return False
    if p == e or p in e or e in p:
        return True
    if p.endswith("シリーズ"):
        base = p[: -len("シリーズ")]
        if base and (e.startswith(base) or base in e):
            return True
    return False


def plane_agrees(plane, keitai):
    k = norm(keitai)
    if plane == "B":
        return "バイ" in k or "IVR" in k
    if plane == "S":
        return "シングル" in k
    return False


def sim(a, b):
    return difflib.SequenceMatcher(None, norm(a), norm(b)).ratio()


def append_cell(ws, row, col, text):
    cur = ws.cell(row, col).value
    ws.cell(row, col).value = f"{cur}、{text}" if cur not in (None, "") else text


def set_red_font(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        f = copy(cell.font)
        f.color = RED
        cell.font = f


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True)
    ap.add_argument("--data", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--sheet", default="まとめ")
    args = ap.parse_args()

    data = json.load(open(args.data, encoding="utf-8"))
    pref, year = data["prefecture"], int(data["year"])
    aliases = data.get("aliases", {})

    wb = openpyxl.load_workbook(args.excel)
    ws = wb[args.sheet]

    # ヘッダー行から列番号を解決(列構成の変化に耐える)
    header = {norm(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1)}
    def col(name):
        c = header.get(norm(name))
        if not c:
            sys.exit(f"ERROR: 列「{name}」がヘッダーに見つかりません")
        return c
    C_HOSP, C_PREF = col("病院名"), col("県")
    C_MAKER, C_YEAR, C_SYS = col("メーカー"), col("納入年度"), col("システム")
    C_TYPE, C_KEITAI, C_PURPOSE = col("TYPE"), col("形態"), col("主目的")
    C_PREV, C_Q, C_BIKO = col("前装置"), col("質問"), col("備考")

    # 対象県の行を抽出し、病院名(原文)ごとにグループ化
    target_rows = [r for r in range(2, ws.max_row + 1)
                   if str(ws.cell(r, C_PREF).value or "").strip() == pref]
    if not target_rows:
        sys.exit(f"ERROR: 県=「{pref}」の行が0件。「県」列の値を確認してください")
    by_hosp = defaultdict(list)
    for r in target_rows:
        by_hosp[str(ws.cell(r, C_HOSP).value)].append(r)
    excel_hosps = list(by_hosp.keys())

    # シート全体からシステム名・メーカー名の既存表記(正規化キー→最頻出の原表記)を学習
    sys_spelling, maker_spelling = Counter(), Counter()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, C_SYS).value
        if v:
            sys_spelling[(norm(v), str(v))] += 1
        m = ws.cell(r, C_MAKER).value
        if m:
            maker_spelling[(maker_key(m), str(m))] += 1
    def canonical_sys(pdf_name):
        cands = [(cnt, raw) for (n, raw), cnt in sys_spelling.items() if n == norm(pdf_name)]
        return max(cands)[1] if cands else str(pdf_name)
    def canonical_maker(abbr):
        k = maker_key(abbr)
        cands = [(cnt, raw) for (mk, raw), cnt in maker_spelling.items() if mk == k]
        return max(cands)[1] if cands else str(abbr)

    # PDF病院 → Excel病院グループ の対応付け
    def match_hospital(pdf_name):
        if pdf_name in aliases:
            tgt = aliases[pdf_name]
            hits = [h for h in by_hosp if norm(h) == norm(tgt)]
        else:
            p = norm(pdf_name)
            hits = [h for h in by_hosp if norm(h).endswith(p) or p in norm(h)]
        return hits

    unchanged, updated, notlisted, pdf_only, new_hosps, warnings = [], [], [], [], [], []
    matched_excel_hosps = set()

    for hosp in data["hospitals"]:
        hits = match_hospital(hosp["name"])
        if len(hits) > 1:
            sys.exit(f"ERROR: PDF病院「{hosp['name']}」がExcelの複数病院 {hits} に一致。"
                     f" aliases で正確な病院名を指定してください")
        if not hits:
            new_hosps.append(hosp)
            continue
        ehosp = hits[0]
        matched_excel_hosps.add(ehosp)
        rows = list(by_hosp[ehosp])          # 未対応のExcel行
        devs = list(hosp["devices"])         # 未対応のPDF装置

        # パス1: 同一システム(ルール1) — メーカー一致も要求
        for d in list(devs):
            for r in rows:
                if maker_key(ws.cell(r, C_MAKER).value) == maker_key(d["maker"]) \
                        and same_system(d["system"], ws.cell(r, C_SYS).value):
                    unchanged.append((r, ehosp, d))
                    rows.remove(r); devs.remove(d)
                    break

        # パス2: 同一メーカー内で名称類似度の高い順にペアリング(装置更新)
        # パス3: 残りをメーカー跨ぎで形態一致優先にペアリング(装置更新・要確認)
        def greedy_pairs(cross_maker):
            pairs = []
            for d in devs:
                for r in rows:
                    same_mk = maker_key(ws.cell(r, C_MAKER).value) == maker_key(d["maker"])
                    if cross_maker == same_mk:
                        continue
                    score = sim(d["system"], ws.cell(r, C_SYS).value) \
                        + (0.5 if plane_agrees(d.get("plane", ""), ws.cell(r, C_KEITAI).value) else 0)
                    pairs.append((score, d, r))
            for score, d, r in sorted(pairs, key=lambda x: -x[0]):
                if d in devs and r in rows:
                    updated.append((r, ehosp, d, cross_maker))
                    devs.remove(d); rows.remove(r)

        greedy_pairs(cross_maker=False)
        greedy_pairs(cross_maker=True)

        for r in rows:                       # ルール3: PDFに無いExcel行
            notlisted.append((r, ehosp))
        for d in devs:                       # PDFにあるがExcelに無い装置
            pdf_only.append((ehosp, by_hosp[ehosp][0], d))

    # PDFに登場しなかったExcel病院 → 全行ルール3
    for ehosp in excel_hosps:
        if ehosp not in matched_excel_hosps:
            for r in by_hosp[ehosp]:
                notlisted.append((r, ehosp))

    # ---- 書き込み ----
    for r, ehosp, d, cross in updated:
        old_maker = ws.cell(r, C_MAKER).value
        old_year = ws.cell(r, C_YEAR).value
        old_sys = ws.cell(r, C_SYS).value
        ws.cell(r, C_MAKER).value = canonical_maker(d["maker"])
        ws.cell(r, C_YEAR).value = year
        ws.cell(r, C_SYS).value = canonical_sys(d["system"])
        ws.cell(r, C_PREV).value = f"{old_year} {old_sys}"
        plane = d.get("plane", "")
        ws.cell(r, C_KEITAI).value = {"B": "バイ", "S": "シングル"}.get(plane, None)
        for c in (C_TYPE, C_PURPOSE) + (() if plane else (C_KEITAI,)):
            ws.cell(r, c).value = None
            ws.cell(r, c).fill = YELLOW
        set_red_font(ws, r, ws.max_column)
        if cross:
            append_cell(ws, r, C_Q, f"{old_maker} {old_sys}→{canonical_maker(d['maker'])} "
                                    f"{canonical_sys(d['system'])}の更新対応で良いか要確認")
        if d.get("mark") == "C":
            append_cell(ws, r, C_Q, f"PDFではⒸ既設表記。納入年度{year}で良いか要確認")

    for r, ehosp in notlisted:
        append_cell(ws, r, C_BIKO, f"{year}年リスト未掲載")

    for ehosp, r, d in pdf_only:
        append_cell(ws, r, C_Q, f"PDFに{canonical_maker(d['maker'])} {d['system']}"
                                f"[{d.get('plane','')}]の記載あり、Excelに対応行なし(要確認)")

    wb.save(args.out)

    # ---- レポート ----
    def show(r):
        return (f"行{r}: {ws.cell(r, C_HOSP).value} | {ws.cell(r, C_MAKER).value} | "
                f"{ws.cell(r, C_YEAR).value} | {ws.cell(r, C_SYS).value}")
    print(f"=== 突合結果 ({pref} / {year}年リスト) ===")
    print(f"対象行: {len(target_rows)}  変更なし: {len(unchanged)}  更新: {len(updated)}  "
          f"未掲載: {len(notlisted)}")
    assert len(target_rows) == len(unchanged) + len(updated) + len(notlisted), "行数が合いません"
    print("\n--- 更新した行(赤文字) ---")
    for r, ehosp, d, cross in updated:
        print(("[要確認] " if cross else "") + show(r) + f"  (前装置: {ws.cell(r, C_PREV).value})")
    print("\n--- 未掲載とした行(備考追記) ---")
    for r, ehosp in notlisted:
        print(show(r))
    print("\n--- PDFにあるがExcelに行が無い装置(質問列に記載済) ---")
    for ehosp, r, d in pdf_only:
        print(f"{ehosp}: {d['maker']} {d['system']} [{d.get('plane','')}] "
              f"マーク={d.get('mark','')}")
    print("\n--- PDFにあるがExcelに無い病院(行追加はしていない。必要ならユーザーに確認) ---")
    for h in new_hosps:
        for d in h["devices"]:
            print(f"{h['name']}: {d['maker']} {d['system']} [{d.get('plane','')}] "
                  f"マーク={d.get('mark','')}")
    print(f"\n保存: {args.out}")


if __name__ == "__main__":
    main()

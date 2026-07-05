# -*- coding: utf-8 -*-
"""施設名簿の年次比較Excelを生成するテンプレート。

データファイル(Python)に rows / 設定を定義して呼び出す:

    from build_excel import build_workbook, SAME

    rows = [
        # (施設名, 旧年版装置, 新年版装置, 区分, 備考)
        # 装置 = [(メーカー, "システム名 [S]"), ...] / 新年版が同一なら SAME
        ("○○病院",
         [("キヤノン", "INFX-8000C [S]"), ("島津", "CVS package [S]")],
         [("島津", "CVS package [S]")],
         "変更", "キヤノン INFX-8000C が新年版でなくなる"),
        ("△△病院", [("GE", "IGS 530 [S]")], SAME, "", ""),
        ("□□病院", [], [("島津", "Trinias シリーズ [S]")], "新規", "新年版で新規掲載"),
    ]

    build_workbook(
        rows,
        out="比較.xlsx",
        sheet="北海道",
        title="北海道 血管造影システム設置施設 メーカー・システム比較(2024年版/2025年版)",
        source="出典:『新医療』設置施設名簿 2024年6月号/2025年6月号。[S]=シングルプレーン,[B]=バイプレーン",
        old_label="2024年版(2024年3月1日現在)",
        new_label="2025年版(2025年3月1日現在)",
    )

区分: "変更"(黄)/ "新規"(緑)/ "掲載なし"(赤)/ ""(色なし)
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SAME = "__SAME__"

FILL_CHANGE = PatternFill("solid", fgColor="FFE699")   # 黄: 入替・増設・減少
FILL_NEW = PatternFill("solid", fgColor="C6E0B4")      # 緑: 新規掲載
FILL_REMOVED = PatternFill("solid", fgColor="F4B8B8")  # 赤: 掲載なし
FILL_HEAD = PatternFill("solid", fgColor="305496")
FILL_HEAD2 = PatternFill("solid", fgColor="8EA9DB")

STATUS_FILL = {"変更": FILL_CHANGE, "新規": FILL_NEW, "掲載なし": FILL_REMOVED}


def _fmt(systems):
    if not systems:
        return "―", "(掲載なし)"
    return "\n".join(m for m, _ in systems), "\n".join(s for _, s in systems)


def build_workbook(rows, out, sheet, title, source, old_label, new_label):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet

    thin = Side(style="thin", color="B0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap_top = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = source
    ws["A2"].font = Font(size=9, color="666666")

    legend = [("装置の入替・増設・減少あり", FILL_CHANGE),
              ("新しい年版で新規掲載", FILL_NEW),
              ("新しい年版に掲載なし", FILL_REMOVED)]
    for i, (txt, fill) in enumerate(legend):
        c = ws.cell(row=3, column=2 + i * 2, value=txt)
        c.fill = fill
        c.font = Font(size=9)
        c.border = border

    hr = 5
    ws.cell(row=hr, column=1, value="No.")
    ws.cell(row=hr, column=2, value="施設名")
    ws.cell(row=hr, column=3, value=old_label)
    ws.cell(row=hr, column=5, value=new_label)
    ws.cell(row=hr, column=7, value="変更区分")
    ws.cell(row=hr, column=8, value="備考")
    ws.merge_cells(start_row=hr, start_column=3, end_row=hr, end_column=4)
    ws.merge_cells(start_row=hr, start_column=5, end_row=hr, end_column=6)
    for col in (1, 2, 7, 8):
        ws.merge_cells(start_row=hr, start_column=col, end_row=hr + 1, end_column=col)
    for col, v in {3: "メーカー", 4: "システム名", 5: "メーカー", 6: "システム名"}.items():
        ws.cell(row=hr + 1, column=col, value=v)
    for r in (hr, hr + 1):
        for col in range(1, 9):
            c = ws.cell(row=r, column=col)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = FILL_HEAD if r == hr else FILL_HEAD2
            c.alignment = center
            c.border = border

    r = hr + 2
    for i, (name, old, new, status, note) in enumerate(rows, 1):
        if new == SAME:
            new = old
        m_old, s_old = _fmt(old)
        m_new, s_new = _fmt(new)
        for col, v in enumerate([i, name, m_old, s_old, m_new, s_new, status, note], 1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = border
            c.alignment = wrap_top if col != 1 else Alignment(horizontal="center", vertical="top")
            c.font = Font(size=10)
        fill = STATUS_FILL.get(status)
        if fill:
            for col in range(2, 8):
                ws.cell(row=r, column=col).fill = fill
        r += 1

    for col, w in {1: 5, 2: 30, 3: 12, 4: 28, 5: 12, 6: 28, 7: 10, 8: 46}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{hr + 2}"
    ws.sheet_view.zoomScale = 90

    wb.save(out)
    print("saved", out, "rows:", len(rows))
    return out

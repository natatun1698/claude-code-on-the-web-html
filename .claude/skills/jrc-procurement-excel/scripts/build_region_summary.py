#!/usr/bin/env python3
"""地域まとめExcel（地域サマリー + X線関連該当一覧）を生成する。

hospitals.json の形式（リスト）:
[
  {"pref": "鳥取県", "key": "tottori", "name": "鳥取赤十字病院",
   "note": "任意の固定備考（省略可。省略時は該当有無で自動設定）"},
  ...
]
"key" は work/contracts_<key>.json に対応する。0件の病院も
（例: echo '[]' > work/contracts_<key>.json しておく）必ずリストに含めること。

usage: python3 build_region_summary.py <地域名> <hospitals.json> <output.xlsx> [contracts_dir]
  地域名: シートタイトルに使う（例: "中国地方" "鳥取県・島根県・岡山県・広島県・山口県"）
  contracts_dir: work/contracts_<key>.json を探すディレクトリ（省略時 "work"）
"""
import glob
import json
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from build_excel import categorize, is_maintenance, date_key, style_header

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def main(region_name, hospitals_json, out_path, contracts_dir="work"):
    hospitals = json.load(open(hospitals_json, encoding="utf-8"))

    wb = Workbook()
    ws = wb.active
    ws.title = "地域サマリー"
    ws.cell(row=1, column=1,
            value=f"{region_name} 赤十字病院 随意契約公表情報（X線装置関連）調査サマリー").font = Font(bold=True, size=14)

    headers = ["都道府県", "病院名", "調査した公表データ件数", "X線関連該当件数", "備考", "Excelファイル"]
    border = style_header(ws, 3, headers, fill="C00000")

    all_matched = []
    r = 4
    for h in hospitals:
        pref, key, name = h["pref"], h["key"], h["name"]
        path = os.path.join(contracts_dir, f"contracts_{key}.json")
        data = json.load(open(path, encoding="utf-8")) if os.path.exists(path) else []
        matched = []
        for row in data:
            cat = categorize(row["name"])
            if cat:
                row = dict(row)
                row["category"] = cat
                row["kind"] = "保守" if is_maintenance(row["name"]) else "製品"
                row["pref"] = pref
                row["hospital"] = name
                matched.append(row)
        all_matched.extend(matched)
        note = h.get("note") or ("" if matched else "X線関連契約の該当なし")
        values = [pref, name, len(data), len(matched), note, f"{name}_随意契約_X線関連.xlsx"]
        for ci, v in enumerate(values, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)
        r += 1

    for ci, w in enumerate([12, 28, 20, 16, 32, 46], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A4"

    ws2 = wb.create_sheet("X線関連該当一覧")
    headers2 = ["都道府県", "病院名", "物品などまたは役務の名称", "締結日", "契約金額",
                "区分（装置カテゴリ）", "製品/保守", "相手方", "URL"]
    border2 = style_header(ws2, 1, headers2, fill="404040")
    all_matched.sort(key=lambda x: (x["hospital"], date_key(x["date"])))
    r = 2
    for row in all_matched:
        amount = row["amount"]
        try:
            amount_v = int(re.sub(r"[^0-9]", "", amount)) if re.search(r"\d", amount) else amount
        except ValueError:
            amount_v = amount
        values = [row["pref"], row["hospital"], row["name"], row["date"], amount_v,
                  row["category"], row["kind"], row["counterparty"], row["source_url"]]
        for ci, v in enumerate(values, 1):
            c = ws2.cell(row=r, column=ci, value=v)
            c.border = border2
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if ci == 5 and isinstance(v, int):
                c.number_format = "#,##0"
        r += 1

    for ci, w in enumerate([12, 28, 40, 14, 14, 30, 10, 34, 48], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A2"

    wb.save(out_path)
    print(f"saved {out_path}: {len(all_matched)} matched rows total across {len(hospitals)} hospitals")


if __name__ == "__main__":
    contracts_dir = sys.argv[4] if len(sys.argv) > 4 else "work"
    main(sys.argv[1], sys.argv[2], sys.argv[3], contracts_dir)

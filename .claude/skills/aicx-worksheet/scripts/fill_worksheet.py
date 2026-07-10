#!/usr/bin/env python3
"""AICX認定「AIエージェント導入ワークシート」(Excel) の空欄をケース定義JSONから埋める。

使い方:
    python3 fill_worksheet.py --template <空欄テンプレ.xlsx> --case <ケース定義.json> --out <出力.xlsx>

ケース定義JSONのスキーマは同梱の examples/invoice_agent.json と SKILL.md を参照。
"""
import argparse
import copy
import json

import openpyxl

SHEET_CH01 = "Ch01_任せたい仕事"
SHEET_CH02 = "Ch02_SIPOC_As-Is"
SHEET_CH03 = "Ch03-04_データ・ワークフロー"
SHEET_CH05 = "Ch05_推進体制・責任"
SHEET_CH06 = "Ch06_5D進行チェック"

# Ch01 G列(月間工数)の自動計算式。テンプレは10行目以降にのみ式があるため、
# 例示用だった7〜9行目に書き込む場合は同じ式を補う。
G_FORMULA = (
    '=IF(OR(E{r}="",F{r}=""),"",ROUND(IF(ISNUMBER(E{r}),E{r},'
    'IF(ISNUMBER(SEARCH("毎日",E{r})),20,'
    'IF(ISNUMBER(SEARCH("週",E{r})),VALUE(MID(E{r},FIND("週",E{r})+1,FIND("回",E{r})-FIND("週",E{r})-1))*4,'
    'IF(ISNUMBER(SEARCH("月",E{r})),VALUE(MID(E{r},FIND("月",E{r})+1,FIND("回",E{r})-FIND("月",E{r})-1)),1))))'
    "*F{r}/60,1))"
)
I_FORMULA = '=IF(G{r}="","",ROUND(G{r}*0.6,1))'


def set_cell(ws, coord, value):
    """値を書き込み、複数行テキストは折り返し表示にする。"""
    cell = ws[coord]
    cell.value = value
    if isinstance(value, str) and "\n" in value:
        al = copy.copy(cell.alignment)
        al.wrap_text = True
        if not al.vertical:
            al.vertical = "top"
        cell.alignment = al


def fill_ch01(ws, data):
    for i, row in enumerate(data.get("rows", [])):
        r = 7 + i  # No.1 は7行目
        set_cell(ws, f"B{r}", row.get("dept", ""))
        set_cell(ws, f"C{r}", row.get("task", ""))
        set_cell(ws, f"D{r}", row.get("issue", ""))
        set_cell(ws, f"E{r}", row.get("freq", ""))
        set_cell(ws, f"F{r}", row.get("minutes", ""))
        set_cell(ws, f"H{r}", row.get("effect", ""))
        set_cell(ws, f"J{r}", row.get("priority", ""))
        set_cell(ws, f"K{r}", row.get("memo", ""))
        if ws[f"G{r}"].value is None:
            ws[f"G{r}"] = G_FORMULA.format(r=r)
        if ws[f"I{r}"].value is None:
            ws[f"I{r}"] = I_FORMULA.format(r=r)


def fill_ch02(ws, data):
    set_cell(ws, "C4", data.get("target", ""))  # 対象業務(値はC4:D4結合セル)
    set_cell(ws, "E5", data.get("date", ""))  # 作成日(ラベルE4の直下)
    set_cell(ws, "F5", data.get("author", ""))  # 作成者(ラベルF4の直下)
    sipoc = data.get("sipoc", {})
    set_cell(ws, "B9", sipoc.get("supplier", ""))
    set_cell(ws, "C9", sipoc.get("input", ""))
    set_cell(ws, "D9", sipoc.get("process", ""))
    set_cell(ws, "E9", sipoc.get("output", ""))
    set_cell(ws, "F9", sipoc.get("customer", ""))
    asis = data.get("asis", {})
    # 各キーはB〜F列(受付/確認/判断/出力/備考)に対応する5要素のリスト
    for key, r in (("overview", 18), ("owner", 19), ("tools", 20), ("issues", 21)):
        values = asis.get(key, [])
        for j, col in enumerate("BCDEF"):
            if j < len(values) and values[j]:
                set_cell(ws, f"{col}{r}", values[j])


def fill_ch03(ws, data):
    set_cell(ws, "D4", "対象業務：" + data.get("target", ""))
    set_cell(ws, "F4", "作成日：" + data.get("date", ""))
    set_cell(ws, "G4", "作成者：" + data.get("author", ""))
    for i, row in enumerate(data.get("rows", [])):
        r = 7 + i  # 1起点=7行目 〜 8記録・監査=14行目
        set_cell(ws, f"B{r}", row.get("overview", ""))
        set_cell(ws, f"C{r}", row.get("io", ""))
        set_cell(ws, f"D{r}", row.get("tool", ""))
        set_cell(ws, f"E{r}", row.get("owner", ""))
        set_cell(ws, f"F{r}", row.get("quality", ""))
        set_cell(ws, f"G{r}", row.get("memo", ""))


def fill_ch05(ws, data):
    for i, row in enumerate(data.get("rows", [])):
        r = 7 + i  # タスク1=7行目 〜 タスク7=13行目
        marks = row.get("marks", [])  # C〜I列(7役割)のR/A/C/I
        for j, col in enumerate("CDEFGHI"):
            if j < len(marks) and marks[j]:
                set_cell(ws, f"{col}{r}", marks[j])
        set_cell(ws, f"J{r}", row.get("memo", ""))


def fill_ch06(ws, data):
    set_cell(ws, "B4", data.get("project", ""))
    set_cell(ws, "F4", data.get("target", ""))
    set_cell(ws, "B5", data.get("author", ""))
    set_cell(ws, "D5", data.get("date", ""))
    set_cell(ws, "F5", data.get("final", ""))
    set_cell(ws, "H5", data.get("refs", ""))
    for row, text in data.get("results", {}).items():
        set_cell(ws, f"E{row}", text)  # 確認結果・参照先 / 判定行はGo/Pivot/No-Goと理由
    for row, text in data.get("memos", {}).items():
        set_cell(ws, f"H{row}", text)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--template", required=True)
    ap.add_argument("--case", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    with open(args.case, encoding="utf-8") as f:
        case = json.load(f)

    wb = openpyxl.load_workbook(args.template)
    fill_ch01(wb[SHEET_CH01], case.get("ch01", {}))
    fill_ch02(wb[SHEET_CH02], case.get("ch02", {}))
    fill_ch03(wb[SHEET_CH03], case.get("ch03_04", {}))
    fill_ch05(wb[SHEET_CH05], case.get("ch05", {}))
    fill_ch06(wb[SHEET_CH06], case.get("ch06", {}))
    wb.save(args.out)
    print(f"saved: {args.out}")


if __name__ == "__main__":
    main()

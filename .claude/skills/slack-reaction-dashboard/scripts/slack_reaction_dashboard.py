#!/usr/bin/env python3
"""Slackリアクション分析ダッシュボード生成ツール。

サブコマンド:
  fetch  Slack Web APIからメッセージ+リアクションを取得してraw JSONを保存
         (要: SLACK_TOKEN 環境変数 or --token。slack.comへのネットワーク到達が必要)
  build  raw JSONを分析し、Excelダッシュボード + Notion向けエクスポートを生成
  demo   サンプルデータでraw JSONを生成してbuildまで実行(動作確認・レイアウト確認用)

使用例:
  python slack_reaction_dashboard.py fetch --channel CBHRRSZAP --days 30 --out data/slack_raw.json
  python slack_reaction_dashboard.py build --raw data/slack_raw.json --outdir output
  python slack_reaction_dashboard.py demo --outdir output

依存: requests, openpyxl, matplotlib (pip install requests openpyxl matplotlib)
"""

import argparse
import csv
import json
import math
import os
import random
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone

try:
    from zoneinfo import ZoneInfo
except ImportError:  # pragma: no cover
    ZoneInfo = None

SCHEMA = "slack-reaction-dashboard/raw-v1"
WEEKDAY_LABELS = ["月", "火", "水", "木", "金", "土", "日"]

# dataviz参照パレット(ライトモード)
C_BLUE = "2A78D6"      # 主系列(量の表現)
C_VIOLET = "4A3AA7"    # 対になる第2測度(されたランキング)
C_GRAY = "C3C2B7"      # ベースライン/残余
INK = "0B0B0B"
INK2 = "52514E"
MUTED = "898781"
GRID = "E1E0D9"
SURFACE = "FCFCFB"


def _now_utc():
    return datetime.now(timezone.utc)


def _tz(tzname):
    if ZoneInfo is not None:
        try:
            return ZoneInfo(tzname)
        except Exception:
            pass
    return timezone(timedelta(hours=9))  # フォールバック: JST


# ---------------------------------------------------------------- fetch

def slack_api(session, token, method, params):
    import requests  # noqa: F401
    url = f"https://slack.com/api/{method}"
    while True:
        resp = session.get(url, params=params,
                           headers={"Authorization": f"Bearer {token}"},
                           timeout=30)
        if resp.status_code == 429:
            wait = int(resp.headers.get("Retry-After", "10"))
            print(f"  rate limited, {wait}s待機...", file=sys.stderr)
            time.sleep(wait + 1)
            continue
        data = resp.json()
        if not data.get("ok"):
            raise RuntimeError(f"Slack API {method} failed: {data.get('error')}")
        return data


def cmd_fetch(args):
    import requests
    token = args.token or os.environ.get("SLACK_TOKEN") or os.environ.get("SLACK_BOT_TOKEN")
    if not token:
        sys.exit("エラー: SLACK_TOKEN 環境変数か --token を指定してください")

    session = requests.Session()
    oldest = (_now_utc() - timedelta(days=args.days)).timestamp()

    # チャンネル名
    channel_name = args.channel
    try:
        info = slack_api(session, token, "conversations.info", {"channel": args.channel})
        channel_name = info["channel"].get("name", args.channel)
    except Exception as e:
        print(f"conversations.info skip: {e}", file=sys.stderr)

    # メッセージ本体
    print(f"#{channel_name} の過去{args.days}日分を取得中...", file=sys.stderr)
    messages, cursor = [], None
    while True:
        params = {"channel": args.channel, "oldest": f"{oldest:.6f}", "limit": 200}
        if cursor:
            params["cursor"] = cursor
        data = slack_api(session, token, "conversations.history", params)
        messages.extend(data.get("messages", []))
        cursor = data.get("response_metadata", {}).get("next_cursor")
        if not cursor:
            break

    # スレッド返信(返信にもリアクションが付くため)
    parents = [m for m in messages if m.get("reply_count") and m.get("thread_ts") == m.get("ts")]
    for i, parent in enumerate(parents):
        print(f"  スレッド {i + 1}/{len(parents)}", file=sys.stderr)
        cursor = None
        while True:
            params = {"channel": args.channel, "ts": parent["ts"],
                      "oldest": f"{oldest:.6f}", "limit": 200}
            if cursor:
                params["cursor"] = cursor
            data = slack_api(session, token, "conversations.replies", params)
            for m in data.get("messages", []):
                if m.get("ts") != parent["ts"]:  # 親の重複を除外
                    messages.append(m)
            cursor = data.get("response_metadata", {}).get("next_cursor")
            if not cursor:
                break

    # ユーザー名マップ
    print("ユーザー一覧を取得中...", file=sys.stderr)
    users, cursor = {}, None
    while True:
        params = {"limit": 200}
        if cursor:
            params["cursor"] = cursor
        data = slack_api(session, token, "users.list", params)
        for u in data.get("members", []):
            prof = u.get("profile", {})
            name = prof.get("display_name") or prof.get("real_name") or u.get("name") or u["id"]
            users[u["id"]] = {"name": name, "is_bot": bool(u.get("is_bot"))}
        cursor = data.get("response_metadata", {}).get("next_cursor")
        if not cursor:
            break

    raw = {
        "schema": SCHEMA,
        "channel": args.channel,
        "channel_name": channel_name,
        "days": args.days,
        "fetched_at": _now_utc().isoformat(),
        "oldest": oldest,
        "users": users,
        "messages": [
            {
                "ts": m.get("ts"),
                "user": m.get("user") or m.get("bot_id"),
                "subtype": m.get("subtype"),
                "thread_ts": m.get("thread_ts"),
                "reactions": [
                    {"name": r.get("name"), "users": r.get("users", []),
                     "count": r.get("count", len(r.get("users", [])))}
                    for r in m.get("reactions", [])
                ],
            }
            for m in messages if m.get("ts")
        ],
    }
    os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(raw, f, ensure_ascii=False, indent=1)
    n_react = sum(len(m["reactions"]) for m in raw["messages"])
    print(f"完了: メッセージ{len(raw['messages'])}件 / リアクション種{n_react}件 → {args.out}")


# ---------------------------------------------------------------- demo data

def cmd_demo_raw(days):
    rng = random.Random(42)
    members = [
        ("U01", "田中"), ("U02", "佐藤"), ("U03", "鈴木"), ("U04", "高橋"),
        ("U05", "伊藤"), ("U06", "渡辺"), ("U07", "山本"), ("U08", "中村"),
        ("U09", "小林"), ("U10", "加藤"), ("U11", "吉田"), ("U12", "山田"),
    ]
    emojis = ["+1", "tada", "eyes", "pray", "joy", "clap", "heart",
              "thinking_face", "fire", "bow", "muscle", "sob"]
    # ユーザーごとに投稿頻度・リアクション性向を偏らせる
    post_w = [rng.uniform(0.3, 3.0) for _ in members]
    react_w = [rng.uniform(0.2, 3.5) for _ in members]
    emoji_w = [rng.uniform(0.3, 3.0) for _ in emojis]
    popularity = [rng.uniform(0.4, 2.5) for _ in members]  # されやすさ

    now = _now_utc()
    messages = []
    for d in range(days):
        day = now - timedelta(days=days - 1 - d)
        is_weekend = day.weekday() >= 5
        n_msgs = max(0, int(rng.gauss(4 if is_weekend else 22, 6)))
        for _ in range(n_msgs):
            hour = min(23, max(0, int(rng.gauss(13, 3.5))))
            ts = day.replace(hour=hour, minute=rng.randrange(60),
                             second=rng.randrange(60), microsecond=0).timestamp()
            ai = rng.choices(range(len(members)), weights=post_w)[0]
            author = members[ai][0]
            reactions = []
            expected = popularity[ai] * (1.6 if rng.random() < 0.25 else 0.7)
            n_reactors = min(len(members) - 1, int(rng.expovariate(1 / max(expected, 0.1))))
            if n_reactors > 0:
                pool = [i for i in range(len(members)) if i != ai]
                weights = [react_w[i] for i in pool]
                chosen = set()
                for _ in range(n_reactors):
                    pick = rng.choices(pool, weights=weights)[0]
                    chosen.add(pick)
                by_emoji = defaultdict(list)
                for i in chosen:
                    e = rng.choices(emojis, weights=emoji_w)[0]
                    by_emoji[e].append(members[i][0])
                reactions = [{"name": e, "users": us, "count": len(us)}
                             for e, us in by_emoji.items()]
            messages.append({"ts": f"{ts:.6f}", "user": author, "subtype": None,
                             "thread_ts": None, "reactions": reactions})
    return {
        "schema": SCHEMA,
        "channel": "CBHRRSZAP",
        "channel_name": "demo-channel",
        "days": days,
        "fetched_at": now.isoformat(),
        "oldest": (now - timedelta(days=days)).timestamp(),
        "users": {uid: {"name": name, "is_bot": False} for uid, name in members},
        "messages": messages,
    }


# ---------------------------------------------------------------- analyze

def analyze(raw, tzname="Asia/Tokyo"):
    tz = _tz(tzname)
    users = raw.get("users", {})

    def uname(uid):
        if uid in users:
            return users[uid]["name"]
        return uid or "(不明)"

    given = Counter()          # リアクションした回数
    received = Counter()       # リアクションされた回数
    emoji = Counter()
    daily_react = Counter()    # メッセージ投稿日ベース(APIはリアクション時刻を返さないため)
    daily_msgs = Counter()
    weekday = Counter()
    hour = Counter()
    msgs_by_user = Counter()
    emoji_variety = defaultdict(set)

    n_msgs = 0
    n_msgs_with = 0
    skip_subtypes = {"channel_join", "channel_leave", "bot_add", "bot_remove"}

    for m in raw["messages"]:
        if m.get("subtype") in skip_subtypes:
            continue
        dt = datetime.fromtimestamp(float(m["ts"]), tz)
        dkey = dt.date().isoformat()
        n_msgs += 1
        daily_msgs[dkey] += 1
        author = uname(m.get("user"))
        msgs_by_user[author] += 1

        total_here = 0
        for r in m.get("reactions", []):
            cnt = r.get("count") or len(r.get("users", []))
            total_here += cnt
            emoji[f":{r['name']}:"] += cnt
            received[author] += cnt
            for uid in r.get("users", []):
                given[uname(uid)] += 1
                emoji_variety[uname(uid)].add(r["name"])
        if total_here:
            n_msgs_with += 1
            daily_react[dkey] += total_here
            weekday[dt.weekday()] += total_here
            hour[dt.hour] += total_here

    total_reactions = sum(emoji.values())
    # 期間内の全日付(欠損日は0埋め)
    end = datetime.fromtimestamp(
        max(float(m["ts"]) for m in raw["messages"]), tz).date() if raw["messages"] else datetime.now(tz).date()
    start = end - timedelta(days=raw.get("days", 30) - 1)
    daily = []
    d = start
    while d <= end:
        k = d.isoformat()
        daily.append((k, daily_react.get(k, 0), daily_msgs.get(k, 0)))
        d += timedelta(days=1)

    names = set(given) | set(received)
    user_rows = sorted(
        ({"name": n, "given": given.get(n, 0), "received": received.get(n, 0),
          "messages": msgs_by_user.get(n, 0),
          "emoji_variety": len(emoji_variety.get(n, ()))}
         for n in names),
        key=lambda r: -(r["given"] + r["received"]))

    return {
        "meta": {
            "channel": raw.get("channel"), "channel_name": raw.get("channel_name"),
            "days": raw.get("days"), "period_start": start.isoformat(),
            "period_end": end.isoformat(), "tz": tzname,
            "generated_at": datetime.now(tz).isoformat(),
        },
        "kpi": {
            "messages": n_msgs,
            "messages_with_reactions": n_msgs_with,
            "total_reactions": total_reactions,
            "unique_reactors": len(given),
            "unique_emoji": len(emoji),
            "avg_per_msg": round(total_reactions / n_msgs, 2) if n_msgs else 0,
            "coverage_pct": round(100 * n_msgs_with / n_msgs, 1) if n_msgs else 0,
        },
        "reactors": given.most_common(),
        "receivers": received.most_common(),
        "emoji": emoji.most_common(),
        "daily": daily,
        "weekday": [(WEEKDAY_LABELS[i], weekday.get(i, 0)) for i in range(7)],
        "hour": [(h, hour.get(h, 0)) for h in range(24)],
        "users": user_rows,
    }


# ---------------------------------------------------------------- PNG charts (Notion用)

def build_pngs(agg, outdir):
    import matplotlib
    matplotlib.use("Agg")
    from matplotlib import font_manager, pyplot as plt

    jp = None
    available = {f.name for f in font_manager.fontManager.ttflist}
    for cand in ("Noto Sans CJK JP", "IPAPGothic", "IPAGothic", "TakaoPGothic"):
        if cand in available:
            jp = cand
            break
    if jp:
        plt.rcParams["font.family"] = jp
    plt.rcParams.update({
        "figure.facecolor": f"#{SURFACE}", "axes.facecolor": f"#{SURFACE}",
        "savefig.facecolor": f"#{SURFACE}",
        "text.color": f"#{INK}", "axes.labelcolor": f"#{INK2}",
        "xtick.color": f"#{MUTED}", "ytick.color": f"#{MUTED}",
        "axes.edgecolor": f"#{C_GRAY}", "axes.grid": True,
        "grid.color": f"#{GRID}", "grid.linewidth": 0.8,
        "axes.spines.top": False, "axes.spines.right": False,
        "font.size": 11, "axes.titlesize": 13, "axes.titleweight": "bold",
    })
    os.makedirs(outdir, exist_ok=True)
    files = {}

    def save(fig, name):
        path = os.path.join(outdir, name)
        fig.tight_layout()
        fig.savefig(path, dpi=150)
        plt.close(fig)
        files[name] = path

    def barh_rank(rows, title, color, fname):
        rows = rows[:10][::-1]
        labels = [r[0] for r in rows]
        vals = [r[1] for r in rows]
        fig, ax = plt.subplots(figsize=(7, 4.2))
        bars = ax.barh(labels, vals, color=f"#{color}", height=0.62)
        ax.set_title(title, loc="left", pad=12)
        ax.grid(axis="y", visible=False)
        vmax = max(vals) if vals else 1
        for b, v in zip(bars, vals):
            ax.text(b.get_width() + vmax * 0.015, b.get_y() + b.get_height() / 2,
                    f"{v:,}", va="center", fontsize=10, color=f"#{INK2}")
        ax.set_xlim(0, vmax * 1.12)
        save(fig, fname)

    barh_rank(agg["reactors"], "リアクションした回数 Top10", C_BLUE, "01_reactors.png")
    barh_rank(agg["receivers"], "リアクションされた回数 Top10", C_VIOLET, "02_receivers.png")
    barh_rank(agg["emoji"], "人気絵文字 Top10", C_BLUE, "03_emoji.png")

    # 日別推移
    dates = [d[5:].replace("-", "/") for d, _, _ in agg["daily"]]
    vals = [v for _, v, _ in agg["daily"]]
    fig, ax = plt.subplots(figsize=(9, 3.6))
    ax.plot(range(len(vals)), vals, color=f"#{C_BLUE}", linewidth=2)
    step = max(1, len(dates) // 10)
    ax.set_xticks(range(0, len(dates), step))
    ax.set_xticklabels(dates[::step])
    ax.set_title("日別リアクション数の推移", loc="left", pad=12)
    ax.grid(axis="x", visible=False)
    ax.set_ylim(bottom=0)
    save(fig, "04_daily.png")

    def col_chart(labels, vals, title, fname, figsize=(7, 3.4)):
        fig, ax = plt.subplots(figsize=figsize)
        ax.bar(range(len(vals)), vals, color=f"#{C_BLUE}", width=0.62)
        ax.set_xticks(range(len(labels)))
        ax.set_xticklabels(labels)
        ax.set_title(title, loc="left", pad=12)
        ax.grid(axis="x", visible=False)
        save(fig, fname)

    col_chart([w for w, _ in agg["weekday"]], [v for _, v in agg["weekday"]],
              "曜日別リアクション数", "05_weekday.png")
    col_chart([str(h) for h, _ in agg["hour"]], [v for _, v in agg["hour"]],
              "時間帯別リアクション数", "06_hour.png", figsize=(9, 3.4))

    # した vs された 散布図
    top = agg["users"][:15]
    fig, ax = plt.subplots(figsize=(7, 5))
    xs = [u["given"] for u in top]
    ys = [u["received"] for u in top]
    ax.scatter(xs, ys, s=90, color=f"#{C_BLUE}", edgecolors=f"#{SURFACE}", linewidths=2, zorder=3)
    for u in top:
        ax.annotate(u["name"], (u["given"], u["received"]),
                    xytext=(6, 5), textcoords="offset points",
                    fontsize=9, color=f"#{INK2}")
    lim = max(xs + ys + [1]) * 1.1
    ax.plot([0, lim], [0, lim], color=f"#{GRID}", linewidth=1, zorder=1)
    ax.set_xlim(0, lim)
    ax.set_ylim(0, lim)
    ax.set_xlabel("リアクションした回数")
    ax.set_ylabel("リアクションされた回数")
    ax.set_title("した vs された(上位15名)", loc="left", pad=12)
    save(fig, "07_scatter.png")

    # リアクション有無比率
    k = agg["kpi"]
    fig, ax = plt.subplots(figsize=(4.6, 4.2))
    wedges, _ = ax.pie(
        [k["messages_with_reactions"], max(k["messages"] - k["messages_with_reactions"], 0)],
        colors=[f"#{C_BLUE}", f"#{GRID}"], startangle=90, counterclock=False,
        wedgeprops={"width": 0.38, "edgecolor": f"#{SURFACE}", "linewidth": 2})
    ax.text(0, 0.08, f"{k['coverage_pct']}%", ha="center", va="center",
            fontsize=26, fontweight="bold", color=f"#{INK}")
    ax.text(0, -0.22, "リアクション付きメッセージ", ha="center", va="center",
            fontsize=9, color=f"#{INK2}")
    ax.set_title("メッセージのリアクション率", loc="left", pad=12)
    save(fig, "08_coverage.png")

    return files


# ---------------------------------------------------------------- Excel

def build_excel(agg, path):
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference, ScatterChart, Series
    from openpyxl.chart.marker import Marker
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    meta, kpi = agg["meta"], agg["kpi"]

    f_title = Font(name="Yu Gothic", size=18, bold=True, color=INK)
    f_sub = Font(name="Yu Gothic", size=10, color=INK2)
    f_kpi_v = Font(name="Yu Gothic", size=20, bold=True, color=INK)
    f_kpi_l = Font(name="Yu Gothic", size=9, color=MUTED)
    f_head = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
    f_cell = Font(name="Yu Gothic", size=10, color=INK)
    head_fill = PatternFill("solid", fgColor=INK2)
    thin = Side(style="thin", color=GRID)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---- データシート(フラット表: Notion/CSVインポート互換) ----
    def data_sheet(name, headers, rows):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for c in ws[1]:
            c.font = f_head
            c.fill = head_fill
            c.border = border
        for row in rows:
            ws.append(list(row))
        for r in ws.iter_rows(min_row=2):
            for c in r:
                c.font = f_cell
                c.border = border
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(14, len(str(h)) * 2 + 4)
        ws.freeze_panes = "A2"
        return ws

    ws_gv = data_sheet("D_した", ["順位", "ユーザー", "リアクションした回数"],
                       [(i + 1, n, v) for i, (n, v) in enumerate(agg["reactors"])])
    ws_rc = data_sheet("D_された", ["順位", "ユーザー", "リアクションされた回数"],
                       [(i + 1, n, v) for i, (n, v) in enumerate(agg["receivers"])])
    ws_em = data_sheet("D_絵文字", ["順位", "絵文字", "使用回数"],
                       [(i + 1, n, v) for i, (n, v) in enumerate(agg["emoji"])])
    ws_dl = data_sheet("D_日別", ["日付", "リアクション数", "メッセージ数"], agg["daily"])
    ws_wd = data_sheet("D_曜日", ["曜日", "リアクション数"], agg["weekday"])
    ws_hr = data_sheet("D_時間帯", ["時間帯", "リアクション数"],
                       [(f"{h}時", v) for h, v in agg["hour"]])
    ws_us = data_sheet("D_ユーザー詳細",
                       ["ユーザー", "した回数", "された回数", "投稿数", "絵文字の種類数"],
                       [(u["name"], u["given"], u["received"], u["messages"],
                         u["emoji_variety"]) for u in agg["users"]])
    ws_cv = data_sheet("D_リアクション率", ["区分", "メッセージ数"],
                       [("リアクションあり", kpi["messages_with_reactions"]),
                        ("リアクションなし", max(kpi["messages"] - kpi["messages_with_reactions"], 0))])

    # ---- ダッシュボード ----
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    for col in range(1, 22):
        ws.column_dimensions[get_column_letter(col)].width = 10

    ws["A1"] = f"Slackリアクション分析ダッシュボード  #{meta.get('channel_name') or meta.get('channel')}"
    ws["A1"].font = f_title
    ws["A2"] = (f"期間: {meta['period_start']} 〜 {meta['period_end']} "
                f"({meta.get('days')}日間) / タイムゾーン: {meta['tz']} / 生成: {meta['generated_at'][:16]}")
    ws["A2"].font = f_sub

    kpis = [
        ("総メッセージ数", kpi["messages"]),
        ("総リアクション数", kpi["total_reactions"]),
        ("リアクション率", f"{kpi['coverage_pct']}%"),
        ("平均リアクション/投稿", kpi["avg_per_msg"]),
        ("リアクションした人数", kpi["unique_reactors"]),
        ("絵文字の種類", kpi["unique_emoji"]),
    ]
    kpi_fill = PatternFill("solid", fgColor="F0EFEC")
    for i, (label, value) in enumerate(kpis):
        col = 1 + i * 3
        c1, c2 = get_column_letter(col), get_column_letter(col + 2)
        ws.merge_cells(f"{c1}4:{c2}4")
        ws.merge_cells(f"{c1}5:{c2}5")
        vc, lc = ws[f"{c1}4"], ws[f"{c1}5"]
        vc.value, vc.font = value, f_kpi_v
        lc.value, lc.font = label, f_kpi_l
        vc.alignment = lc.alignment = Alignment(horizontal="center")
        for row in (4, 5):
            for cc in range(col, col + 3):
                ws.cell(row=row, column=cc).fill = kpi_fill
    ws.row_dimensions[4].height = 30

    def style_bar(chart, color):
        chart.style = None
        for s in chart.series:
            s.graphicalProperties.solidFill = color
            s.graphicalProperties.line.noFill = True
        chart.legend = None
        chart.width, chart.height = 17, 9

    def bar_rank(ws_data, n_rows, title, color):
        ch = BarChart()
        ch.type = "bar"
        ch.title = title
        n = min(n_rows, 10)
        # 上位10件・1位が上に来るよう逆順参照はExcel仕様上不可のためカテゴリ軸を反転
        data = Reference(ws_data, min_col=3, min_row=1, max_row=n + 1)
        cats = Reference(ws_data, min_col=2, min_row=2, max_row=n + 1)
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        ch.y_axis.scaling.orientation = "maxMin"
        ch.x_axis.delete = False
        ch.y_axis.delete = False
        style_bar(ch, color)
        return ch

    ws.add_chart(bar_rank(ws_gv, len(agg["reactors"]), "① リアクションした回数 Top10", C_BLUE), "A8")
    ws.add_chart(bar_rank(ws_rc, len(agg["receivers"]), "② リアクションされた回数 Top10", C_VIOLET), "K8")
    ws.add_chart(bar_rank(ws_em, len(agg["emoji"]), "③ 人気絵文字 Top10", C_BLUE), "A27")

    # 日別推移(折れ線)
    ch = LineChart()
    ch.title = "④ 日別リアクション数の推移"
    n = len(agg["daily"])
    ch.add_data(Reference(ws_dl, min_col=2, min_row=1, max_row=n + 1), titles_from_data=True)
    ch.set_categories(Reference(ws_dl, min_col=1, min_row=2, max_row=n + 1))
    ch.series[0].graphicalProperties.line.solidFill = C_BLUE
    ch.series[0].graphicalProperties.line.width = 25000
    ch.series[0].smooth = False
    ch.legend = None
    ch.width, ch.height = 17, 9
    ws.add_chart(ch, "K27")

    # 曜日別・時間帯別(縦棒)
    def col_chart(ws_data, n_rows, title, anchor):
        ch = BarChart()
        ch.type = "col"
        ch.title = title
        ch.add_data(Reference(ws_data, min_col=2, min_row=1, max_row=n_rows + 1),
                    titles_from_data=True)
        ch.set_categories(Reference(ws_data, min_col=1, min_row=2, max_row=n_rows + 1))
        style_bar(ch, C_BLUE)
        ws.add_chart(ch, anchor)

    col_chart(ws_wd, 7, "⑤ 曜日別リアクション数", "A46")
    col_chart(ws_hr, 24, "⑥ 時間帯別リアクション数", "K46")

    # した vs された(散布図)
    ch = ScatterChart()
    ch.title = "⑦ した vs された(ユーザー別)"
    ch.x_axis.title = "リアクションした回数"
    ch.y_axis.title = "リアクションされた回数"
    n = min(len(agg["users"]), 15)
    xref = Reference(ws_us, min_col=2, min_row=2, max_row=n + 1)
    yref = Reference(ws_us, min_col=3, min_row=2, max_row=n + 1)
    s = Series(yref, xref, title="ユーザー")
    s.marker = Marker(symbol="circle", size=8)
    s.marker.graphicalProperties.solidFill = C_BLUE
    s.graphicalProperties.line.noFill = True
    ch.series.append(s)
    ch.legend = None
    ch.x_axis.delete = False
    ch.y_axis.delete = False
    ch.width, ch.height = 17, 9
    ws.add_chart(ch, "A65")

    # リアクション率(ドーナツ)
    ch = DoughnutChart()
    ch.title = "⑧ メッセージのリアクション率"
    ch.add_data(Reference(ws_cv, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    ch.set_categories(Reference(ws_cv, min_col=1, min_row=2, max_row=3))
    ch.holeSize = 55
    from openpyxl.chart.series import DataPoint
    pts = [DataPoint(idx=0), DataPoint(idx=1)]
    pts[0].graphicalProperties.solidFill = C_BLUE
    pts[1].graphicalProperties.solidFill = GRID
    ch.series[0].data_points = pts
    ch.width, ch.height = 17, 9
    ws.add_chart(ch, "K65")

    # ---- READMEシート ----
    ws_r = wb.create_sheet("README")
    notes = [
        "このブックについて",
        "",
        f"対象: Slackチャンネル #{meta.get('channel_name') or meta.get('channel')} ({meta.get('channel')})",
        f"期間: {meta['period_start']} 〜 {meta['period_end']} ({meta.get('days')}日間)",
        "",
        "シート構成:",
        "  Dashboard … KPIと8つのグラフ",
        "  D_* … 各グラフの元データ(フラット表)。CSV/Notionデータベースへそのままインポート可能",
        "",
        "注意事項:",
        "  ・日別/曜日別/時間帯別の集計は「リアクションが付いたメッセージの投稿日時」ベースです",
        "    (Slack APIはリアクション自体の時刻を返さないため)",
        "  ・「した回数」は ユーザー×絵文字×メッセージ の組み合わせを1回と数えます",
        "",
        "Notionへの掲載について:",
        "  ・Excelのグラフはそのまま貼れないため、同梱の notion_export/ フォルダを使ってください",
        "    - dashboard.md … そのままNotionにインポートできるMarkdown",
        "    - *.png … 各グラフ画像(Notionページに埋め込み用)",
        "    - *.csv … 各ランキング表(Notionデータベースとしてインポート用)",
    ]
    for i, line in enumerate(notes, 1):
        c = ws_r.cell(row=i, column=1, value=line)
        c.font = Font(name="Yu Gothic", size=14 if i == 1 else 10,
                      bold=(i == 1), color=INK)
    ws_r.column_dimensions["A"].width = 100

    os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
    wb.save(path)


# ---------------------------------------------------------------- Notion export

def build_notion_export(agg, outdir, png_files):
    os.makedirs(outdir, exist_ok=True)
    meta, kpi = agg["meta"], agg["kpi"]

    def write_csv(name, headers, rows):
        with open(os.path.join(outdir, name), "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(headers)
            w.writerows(rows)

    write_csv("reactors.csv", ["順位", "ユーザー", "リアクションした回数"],
              [(i + 1, n, v) for i, (n, v) in enumerate(agg["reactors"])])
    write_csv("receivers.csv", ["順位", "ユーザー", "リアクションされた回数"],
              [(i + 1, n, v) for i, (n, v) in enumerate(agg["receivers"])])
    write_csv("emoji.csv", ["順位", "絵文字", "使用回数"],
              [(i + 1, n, v) for i, (n, v) in enumerate(agg["emoji"])])
    write_csv("daily.csv", ["日付", "リアクション数", "メッセージ数"], agg["daily"])
    write_csv("users.csv", ["ユーザー", "した回数", "された回数", "投稿数", "絵文字の種類数"],
              [(u["name"], u["given"], u["received"], u["messages"], u["emoji_variety"])
               for u in agg["users"]])

    def md_table(headers, rows):
        out = ["| " + " | ".join(headers) + " |",
               "|" + "|".join(["---"] * len(headers)) + "|"]
        out += ["| " + " | ".join(str(c) for c in r) + " |" for r in rows]
        return "\n".join(out)

    lines = [
        f"# Slackリアクション分析 #{meta.get('channel_name') or meta.get('channel')}",
        "",
        f"**期間**: {meta['period_start']} 〜 {meta['period_end']}({meta.get('days')}日間) "
        f"/ **生成**: {meta['generated_at'][:16]}",
        "",
        "## サマリー",
        "",
        md_table(["指標", "値"], [
            ("総メッセージ数", f"{kpi['messages']:,}"),
            ("総リアクション数", f"{kpi['total_reactions']:,}"),
            ("リアクション率", f"{kpi['coverage_pct']}%"),
            ("平均リアクション/投稿", kpi["avg_per_msg"]),
            ("リアクションした人数", kpi["unique_reactors"]),
            ("絵文字の種類", kpi["unique_emoji"]),
        ]),
        "",
        "## ① リアクションした回数ランキング",
        "",
        "![リアクションした回数 Top10](01_reactors.png)",
        "",
        md_table(["順位", "ユーザー", "回数"],
                 [(i + 1, n, v) for i, (n, v) in enumerate(agg["reactors"][:10])]),
        "",
        "## ② リアクションされた回数ランキング",
        "",
        "![リアクションされた回数 Top10](02_receivers.png)",
        "",
        md_table(["順位", "ユーザー", "回数"],
                 [(i + 1, n, v) for i, (n, v) in enumerate(agg["receivers"][:10])]),
        "",
        "## ③ 人気絵文字 Top10",
        "",
        "![人気絵文字](03_emoji.png)",
        "",
        "## ④ 日別リアクション数の推移",
        "",
        "![日別推移](04_daily.png)",
        "",
        "## ⑤ 曜日別 / ⑥ 時間帯別",
        "",
        "![曜日別](05_weekday.png)",
        "",
        "![時間帯別](06_hour.png)",
        "",
        "## ⑦ した vs された",
        "",
        "![した vs された](07_scatter.png)",
        "",
        "## ⑧ メッセージのリアクション率",
        "",
        "![リアクション率](08_coverage.png)",
        "",
        "---",
        "",
        "> 日別/曜日別/時間帯別はメッセージ投稿日時ベースの集計です"
        "(Slack APIはリアクション時刻を返さないため)。",
    ]
    with open(os.path.join(outdir, "dashboard.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")


# ---------------------------------------------------------------- commands

def run_build(raw, outdir, tzname):
    agg = analyze(raw, tzname)
    os.makedirs(outdir, exist_ok=True)
    with open(os.path.join(outdir, "aggregates.json"), "w", encoding="utf-8") as f:
        json.dump(agg, f, ensure_ascii=False, indent=1)
    notion_dir = os.path.join(outdir, "notion_export")
    pngs = build_pngs(agg, notion_dir)
    build_notion_export(agg, notion_dir, pngs)
    xlsx = os.path.join(outdir, "slack_reaction_dashboard.xlsx")
    build_excel(agg, xlsx)
    k = agg["kpi"]
    print(f"生成完了: {xlsx}")
    print(f"  Notion用: {notion_dir}/ (dashboard.md + PNG x8 + CSV x5)")
    print(f"  メッセージ{k['messages']}件 / リアクション{k['total_reactions']}件 / "
          f"リアクション率{k['coverage_pct']}%")


def cmd_build(args):
    with open(args.raw, encoding="utf-8") as f:
        raw = json.load(f)
    if raw.get("schema") != SCHEMA:
        print(f"警告: schemaが{SCHEMA}ではありません", file=sys.stderr)
    run_build(raw, args.outdir, args.tz)


def cmd_demo(args):
    raw = cmd_demo_raw(args.days)
    os.makedirs(args.outdir, exist_ok=True)
    with open(os.path.join(args.outdir, "slack_raw_demo.json"), "w", encoding="utf-8") as f:
        json.dump(raw, f, ensure_ascii=False)
    run_build(raw, args.outdir, args.tz)


def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = p.add_subparsers(dest="cmd", required=True)

    pf = sub.add_parser("fetch", help="Slack APIからデータ取得")
    pf.add_argument("--channel", required=True, help="チャンネルID (例: CBHRRSZAP)")
    pf.add_argument("--days", type=int, default=30)
    pf.add_argument("--out", default="data/slack_raw.json")
    pf.add_argument("--token", default=None, help="未指定時は$SLACK_TOKEN")
    pf.set_defaults(func=cmd_fetch)

    pb = sub.add_parser("build", help="raw JSONからダッシュボード生成")
    pb.add_argument("--raw", required=True)
    pb.add_argument("--outdir", default="output")
    pb.add_argument("--tz", default="Asia/Tokyo")
    pb.set_defaults(func=cmd_build)

    pd = sub.add_parser("demo", help="サンプルデータで一気通貫実行")
    pd.add_argument("--days", type=int, default=30)
    pd.add_argument("--outdir", default="output")
    pd.add_argument("--tz", default="Asia/Tokyo")
    pd.set_defaults(func=cmd_demo)

    args = p.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()

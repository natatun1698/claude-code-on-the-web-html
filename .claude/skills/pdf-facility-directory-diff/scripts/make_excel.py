#!/usr/bin/env python3
"""Generate the deliverable Excel from verified diff results.

Input: verified.json = [{pref, hospital, maker, model, year, note?}]
Output: xlsx with columns 都道府県 / 病院名 / メーカー名 / システム名 / 予想される導入年
"""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PREF_ORDER = ['北海道', '青森県', '岩手県', '宮城県', '秋田県', '山形県', '福島県',
              '茨城県', '栃木県', '群馬県', '埼玉県', '千葉県', '東京都', '神奈川県']


def build(rows, out_path, title_note):
    wb = Workbook()
    ws = wb.active
    ws.title = '更新システム一覧'
    headers = ['都道府県', '病院名', 'メーカー名', 'システム名', '予想される導入年']
    ws.append(headers)
    hfill = PatternFill('solid', fgColor='1F4E79')
    hfont = Font(color='FFFFFF', bold=True)
    thin = Border(*[Side(style='thin', color='BFBFBF')] * 4)
    for c in ws[1]:
        c.fill, c.font = hfill, hfont
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin
    rows = sorted(rows, key=lambda r: (PREF_ORDER.index(r['pref']) if r['pref'] in PREF_ORDER else 99,
                                       r.get('order', 0)))
    for r in rows:
        ws.append([r['pref'], r['hospital'], r['maker'], r['model'], r['year']])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.border = thin
            c.alignment = Alignment(vertical='center')
    widths = [10, 34, 14, 30, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:E{ws.max_row}'

    info = wb.create_sheet('注記')
    for line in title_note:
        info.append([line])
    info.column_dimensions['A'].width = 110
    wb.save(out_path)
    print('saved', out_path, 'rows:', len(rows))


if __name__ == '__main__':
    rows = json.load(open(sys.argv[1]))
    note = [
        '本一覧は「新医療」誌の血管造影システム設置施設名簿の2024年版(2024年3月1日現在)と',
        '2025年版(2025年3月1日現在)を比較し、2025年版で新たに掲載された(=1年の間に更新・新規導入されたと',
        '推定される)システムを抽出したものです。',
        '',
        '・対象範囲: 北海道〜神奈川県(両年版が共にカバーする範囲。2025年版は新潟県以降未掲載のため)',
        '・「予想される導入年」: 2024年3月〜2025年2月の間に導入されたと推定されます。',
        '・スキャンPDFのOCR＋画像照合により作成。機種名の細部(オプション表記等)は原本をご確認ください。',
    ]
    build(rows, sys.argv[2] if len(sys.argv) > 2 else 'CVS更新一覧_2024-2025.xlsx', note)

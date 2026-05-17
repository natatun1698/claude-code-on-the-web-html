"""
Slack リアクション分析ツール
============================
指定チャンネルの過去30日間のリアクションを集計し、
Excelレポートを生成してGmailで送信します。
"""

import os
import time
import logging
from datetime import datetime, timedelta, timezone
from collections import defaultdict

import requests
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ログ設定（INFOレベルでコンソールに出力）
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# .env ファイルから環境変数を読み込む
load_dotenv()

# ==============================
# 設定値
# ==============================
SLACK_BOT_TOKEN = os.environ.get("SLACK_BOT_TOKEN", "")
CHANNEL_ID = "CBHRRSZAP"          # 分析対象チャンネル
DAYS_TO_ANALYZE = 30              # 集計対象日数
OUTPUT_DIR = "output"             # 出力ディレクトリ
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "slack_reaction_report.xlsx")

# Slack API レート制限対策：リクエスト間の待機秒数
API_SLEEP_SEC = 1.2

# Slack API のベースURL
SLACK_API_BASE = "https://slack.com/api"


# ==============================
# Slack API クライアント
# ==============================

def slack_get(method: str, params: dict) -> dict:
    """
    Slack Web API への GETリクエストを送信する汎用関数。
    レート制限（HTTP 429）が返った場合は Retry-After を見て待機する。

    Args:
        method: API メソッド名（例: "conversations.history"）
        params: クエリパラメータの辞書

    Returns:
        API レスポンスの JSON を dict として返す

    Raises:
        RuntimeError: API エラーが発生した場合
    """
    url = f"{SLACK_API_BASE}/{method}"
    headers = {"Authorization": f"Bearer {SLACK_BOT_TOKEN}"}

    for attempt in range(5):  # 最大5回リトライ
        response = requests.get(url, headers=headers, params=params, timeout=30)

        # レート制限に引っかかった場合
        if response.status_code == 429:
            retry_after = int(response.headers.get("Retry-After", 30))
            logger.warning("Rate limited. Waiting %d seconds...", retry_after)
            time.sleep(retry_after)
            continue

        response.raise_for_status()
        data = response.json()

        # Slack API は HTTP 200 でもエラーを返すことがある
        if not data.get("ok"):
            error_msg = data.get("error", "unknown_error")
            raise RuntimeError(f"Slack API error [{method}]: {error_msg}")

        return data

    raise RuntimeError(f"Slack API [{method}] failed after 5 retries")


# ==============================
# ユーザー情報の取得
# ==============================

def fetch_users() -> dict:
    """
    Slack ワークスペースの全ユーザー情報を取得し、
    ユーザーID → 表示名 のマッピング辞書を返す。
    ボットユーザーも含めて取得し、後でフィルタリングできるように情報を持つ。

    Returns:
        {
            "UXXXXXXX": {
                "name": "yamada.taro",
                "display_name": "山田太郎",
                "real_name": "Yamada Taro",
                "is_bot": False,
            },
            ...
        }
    """
    logger.info("Fetching user list...")
    users = {}
    cursor = None  # ページネーション用カーソル

    while True:
        params = {"limit": 200}
        if cursor:
            params["cursor"] = cursor

        data = slack_get("users.list", params)
        members = data.get("members", [])

        for member in members:
            uid = member["id"]
            profile = member.get("profile", {})

            # 表示名の優先順位: display_name > real_name > name
            display_name = (
                profile.get("display_name")
                or profile.get("real_name")
                or member.get("name", uid)
            )

            users[uid] = {
                "name": member.get("name", uid),
                "display_name": display_name,
                "real_name": profile.get("real_name", ""),
                "is_bot": member.get("is_bot", False),
                "is_app_user": member.get("is_app_user", False),
            }

        # 次ページがあるか確認
        next_cursor = data.get("response_metadata", {}).get("next_cursor", "")
        if not next_cursor:
            break  # 全ページ取得完了
        cursor = next_cursor
        time.sleep(API_SLEEP_SEC)

    logger.info("Fetched %d users.", len(users))
    return users


def get_display_name(users: dict, uid: str) -> str:
    """ユーザーIDから表示名を取得する。不明な場合はIDをそのまま返す。"""
    if uid in users:
        return users[uid]["display_name"] or users[uid]["name"] or uid
    return uid


def is_bot(users: dict, uid: str) -> bool:
    """指定ユーザーがボットかどうか判定する。"""
    if uid not in users:
        return False
    user = users[uid]
    return user.get("is_bot", False) or user.get("is_app_user", False)


# ==============================
# メッセージの取得
# ==============================

def fetch_messages(oldest_ts: float) -> list:
    """
    チャンネルのメッセージをページネーションで全件取得する。
    スレッドの返信も含めて取得する。

    Args:
        oldest_ts: この UNIX タイムスタンプ以降のメッセージを取得

    Returns:
        メッセージオブジェクトのリスト（スレッド返信を含む）
    """
    logger.info("Fetching messages from channel %s ...", CHANNEL_ID)
    all_messages = []
    cursor = None

    # --- チャンネルの履歴を取得 ---
    while True:
        params = {
            "channel": CHANNEL_ID,
            "oldest": str(oldest_ts),
            "limit": 200,  # 1回のAPIコールで取得する最大件数
        }
        if cursor:
            params["cursor"] = cursor

        data = slack_get("conversations.history", params)
        messages = data.get("messages", [])
        all_messages.extend(messages)

        logger.info("  Fetched %d messages so far...", len(all_messages))

        # スレッド返信を取得する（reply_count > 0 のメッセージ）
        for msg in messages:
            if msg.get("reply_count", 0) > 0:
                thread_ts = msg.get("ts")
                all_messages.extend(
                    fetch_thread_replies(thread_ts, oldest_ts)
                )

        # 次ページの確認
        next_cursor = data.get("response_metadata", {}).get("next_cursor", "")
        if not next_cursor:
            break
        cursor = next_cursor
        time.sleep(API_SLEEP_SEC)

    logger.info("Total messages (including thread replies): %d", len(all_messages))
    return all_messages


def fetch_thread_replies(thread_ts: str, oldest_ts: float) -> list:
    """
    スレッドの返信メッセージを取得する。

    Args:
        thread_ts: スレッドの親メッセージのタイムスタンプ
        oldest_ts: この日時より前の返信は除外

    Returns:
        スレッド内の返信メッセージのリスト（親メッセージを除く）
    """
    replies = []
    cursor = None

    while True:
        params = {
            "channel": CHANNEL_ID,
            "ts": thread_ts,
            "oldest": str(oldest_ts),
            "limit": 200,
        }
        if cursor:
            params["cursor"] = cursor

        try:
            data = slack_get("conversations.replies", params)
        except RuntimeError as e:
            logger.warning("Failed to fetch thread replies for ts=%s: %s", thread_ts, e)
            break

        msgs = data.get("messages", [])
        # 最初のメッセージは親なので除外（インデックス1以降が返信）
        replies.extend(msgs[1:])

        next_cursor = data.get("response_metadata", {}).get("next_cursor", "")
        if not next_cursor:
            break
        cursor = next_cursor
        time.sleep(API_SLEEP_SEC)

    return replies


# ==============================
# リアクションの集計
# ==============================

def analyze_reactions(messages: list, users: dict) -> tuple:
    """
    メッセージリストからリアクション情報を集計する。

    Args:
        messages: Slackメッセージオブジェクトのリスト
        users: ユーザーID → ユーザー情報 のマッピング

    Returns:
        (reactions_made, reactions_received, emoji_total)

        reactions_made[user_id]  = {emoji: count, ...}
            → そのユーザーが各絵文字でリアクションした回数

        reactions_received[user_id] = {emoji: count, ...}
            → そのユーザーの投稿が各絵文字でリアクションされた回数

        emoji_total[emoji] = count
            → 絵文字ごとの総リアクション数
    """
    # ユーザーがリアクションした回数: {user_id: {emoji: count}}
    reactions_made = defaultdict(lambda: defaultdict(int))

    # ユーザーの投稿がリアクションされた回数: {user_id: {emoji: count}}
    reactions_received = defaultdict(lambda: defaultdict(int))

    # 絵文字ごとの総使用回数: {emoji: count}
    emoji_total = defaultdict(int)

    for msg in messages:
        poster_id = msg.get("user") or msg.get("bot_id")

        # ボット投稿は集計対象外
        if not poster_id:
            continue
        if msg.get("subtype") == "bot_message":
            continue
        if is_bot(users, poster_id):
            continue

        # このメッセージに付いているリアクションを処理
        for reaction in msg.get("reactions", []):
            emoji = reaction.get("name", "unknown")
            reactor_ids = reaction.get("users", [])

            for reactor_id in reactor_ids:
                # ボットによるリアクションは除外
                if is_bot(users, reactor_id):
                    continue

                # リアクションした人の集計
                reactions_made[reactor_id][emoji] += 1

                # リアクションされた投稿者の集計
                reactions_received[poster_id][emoji] += 1

                # 絵文字全体の集計
                emoji_total[emoji] += 1

    return reactions_made, reactions_received, emoji_total


# ==============================
# Excel レポートの生成
# ==============================

# Excel のヘッダー行に使うスタイル
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="4472C4")  # 青色
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def style_header_row(ws, row_num: int, col_count: int):
    """指定した行をヘッダースタイルで装飾する。"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def auto_column_width(ws):
    """全カラムの幅をコンテンツに合わせて自動調整する。"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def build_most_reactive_users_sheet(ws, reactions_made: dict, users: dict):
    """
    シート1: most_reactive_users
    リアクションした回数が多いユーザーのランキングを作成する。

    列: rank | user_name | reactions_made | favorite_emoji
    """
    headers = ["rank", "user_name", "reactions_made", "favorite_emoji"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    # ユーザーごとにリアクション合計を計算し、降順ソート
    user_totals = []
    for uid, emoji_counts in reactions_made.items():
        total = sum(emoji_counts.values())
        # 最も多く使った絵文字（favorite_emoji）を取得
        fav_emoji = max(emoji_counts, key=emoji_counts.get) if emoji_counts else ""
        user_totals.append((uid, total, fav_emoji))

    user_totals.sort(key=lambda x: x[1], reverse=True)

    for rank, (uid, total, fav_emoji) in enumerate(user_totals, start=1):
        name = get_display_name(users, uid)
        ws.append([rank, name, total, f":{fav_emoji}:"])

    auto_column_width(ws)


def build_most_reacted_users_sheet(ws, reactions_received: dict, users: dict):
    """
    シート2: most_reacted_users
    リアクションされた回数が多いユーザーのランキングを作成する。

    列: rank | user_name | reactions_received | top_received_emoji
    """
    headers = ["rank", "user_name", "reactions_received", "top_received_emoji"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    # ユーザーごとに受け取ったリアクション合計を計算し、降順ソート
    user_totals = []
    for uid, emoji_counts in reactions_received.items():
        total = sum(emoji_counts.values())
        top_emoji = max(emoji_counts, key=emoji_counts.get) if emoji_counts else ""
        user_totals.append((uid, total, top_emoji))

    user_totals.sort(key=lambda x: x[1], reverse=True)

    for rank, (uid, total, top_emoji) in enumerate(user_totals, start=1):
        name = get_display_name(users, uid)
        ws.append([rank, name, total, f":{top_emoji}:"])

    auto_column_width(ws)


def build_emoji_ranking_sheet(ws, emoji_total: dict):
    """
    シート3: emoji_ranking
    絵文字ごとの使用回数ランキングを作成する。

    列: emoji | count
    """
    headers = ["emoji", "count"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    # 使用回数の多い順にソート
    sorted_emojis = sorted(emoji_total.items(), key=lambda x: x[1], reverse=True)

    for emoji, count in sorted_emojis:
        ws.append([f":{emoji}:", count])

    auto_column_width(ws)


def generate_excel_report(
    reactions_made: dict,
    reactions_received: dict,
    emoji_total: dict,
    users: dict,
) -> str:
    """
    Excel レポートを生成してファイルに保存する。

    Returns:
        保存したファイルパス
    """
    logger.info("Generating Excel report...")

    # 出力ディレクトリが無ければ作成
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = openpyxl.Workbook()

    # デフォルトシートを削除して、名前付きシートを追加
    wb.remove(wb.active)

    ws1 = wb.create_sheet("most_reactive_users")
    build_most_reactive_users_sheet(ws1, reactions_made, users)

    ws2 = wb.create_sheet("most_reacted_users")
    build_most_reacted_users_sheet(ws2, reactions_received, users)

    ws3 = wb.create_sheet("emoji_ranking")
    build_emoji_ranking_sheet(ws3, emoji_total)

    wb.save(OUTPUT_FILE)
    logger.info("Excel report saved: %s", OUTPUT_FILE)
    return OUTPUT_FILE


# ==============================
# メイン処理
# ==============================

def main():
    """メイン処理: データ取得 → 集計 → レポート生成 → メール送信"""

    # --- 事前チェック ---
    if not SLACK_BOT_TOKEN:
        raise ValueError("SLACK_BOT_TOKEN が設定されていません。.env を確認してください。")

    # 集計対象の開始日時（現在から30日前）
    now = datetime.now(timezone.utc)
    oldest_dt = now - timedelta(days=DAYS_TO_ANALYZE)
    oldest_ts = oldest_dt.timestamp()

    logger.info("Analysis period: %s → %s", oldest_dt.strftime("%Y-%m-%d"), now.strftime("%Y-%m-%d"))

    # --- Step 1: ユーザー一覧を取得 ---
    users = fetch_users()

    # --- Step 2: チャンネルのメッセージを取得（スレッド返信含む）---
    messages = fetch_messages(oldest_ts)

    if not messages:
        logger.warning("No messages found in the specified period.")

    # --- Step 3: リアクション集計 ---
    logger.info("Analyzing reactions...")
    reactions_made, reactions_received, emoji_total = analyze_reactions(messages, users)

    logger.info("Unique reactors: %d", len(reactions_made))
    logger.info("Unique posters with reactions: %d", len(reactions_received))
    logger.info("Unique emojis used: %d", len(emoji_total))

    # --- Step 4: Excel レポート生成 ---
    report_date = now.strftime("%Y-%m-%d")
    output_path = generate_excel_report(reactions_made, reactions_received, emoji_total, users)

    logger.info("Done! Report: %s", output_path)


if __name__ == "__main__":
    main()

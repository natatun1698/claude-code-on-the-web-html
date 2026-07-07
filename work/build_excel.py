#!/usr/bin/env python3
"""随意契約公表データからX線装置関連契約のExcelを生成する。

usage: python3 build_excel.py <病院名> <contracts.json...> <output.xlsx>
"""
import json
import os
import re
import sys
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 対象装置カテゴリと判定キーワード（NFKC正規化後に照合）
CATEGORIES = [
    ("一般撮影（レントゲン）",
     ["一般撮影", "レントゲン", "X線撮影装置", "デジタルラジオグラフィ",
      "DR装置", "FPD", "FCR", "画像読み取り装置", "画像読取装置", "撮影台"]),
    ("透視撮影台（X線テレビ）",
     ["X線テレビ", "X線TV", "透視撮影台", "据置型X線透視", "デジタル透視"]),
    ("血管撮影（CVS、アンギオ）",
     ["血管撮影", "アンギオ", "血管造影", "循環器X線", "心血管X線",
      "バイプレーン", "CVS"]),
    ("外科用イメージ（可搬型Cアーム透視装置）",
     ["外科用イメージ", "外科用X線", "Cアーム", "移動型汎用X線透視",
      "移動型X線透視", "可搬型透視"]),
    ("回診用X線装置",
     ["回診用", "回診車", "移動型X線撮影", "ポータブル撮影",
      "移動型汎用一体型X線"]),
]

MAINTENANCE_KW = ["保守", "点検", "メンテナンス", "修理"]

# 病院ごとの注記。BUILD_EXCEL_NOTES 環境変数でテキストファイル（1行1注記）を指定すると読み込む。
DEFAULT_NOTES = [
    "・公表対象は日本赤十字社会計規則に基づく随意契約のみ。競争入札による調達は含まれない。",
]


def load_extra_notes():
    path = os.environ.get("BUILD_EXCEL_NOTES")
    if path and os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            return [line.rstrip("\n") for line in f if line.strip()]
    return list(DEFAULT_NOTES)


EXTRA_NOTES = load_extra_notes()
NOTES = {}


def norm(s):
    return unicodedata.normalize("NFKC", s or "")


def date_key(s):
    m = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", norm(s))
    return (int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else (9999, 0, 0)


def fiscal_year(s):
    y, m, _ = date_key(s)
    if y == 9999:
        return "不明"
    return f"{y if m >= 4 else y - 1}年度"


def categorize(name):
    n = norm(name)
    for cat, kws in CATEGORIES:
        if any(norm(k) in n for k in kws):
            return cat
    return None


def is_maintenance(name):
    n = norm(name)
    return any(norm(k) in n for k in MAINTENANCE_KW)


def style_header(ws, row, cols, fill="C00000"):
    head_fill = PatternFill("solid", fgColor=fill)
    head_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = head_fill
        c.font = head_font
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return border


def main(hospital, json_paths, out_path):
    rows = []
    for p in json_paths:
        rows.extend(json.load(open(p)))

    matched = []
    for row in rows:
        cat = categorize(row["name"])
        if cat:
            row["category"] = cat
            row["kind"] = "保守" if is_maintenance(row["name"]) else "製品"
            row["note"] = NOTES.get(row["name"], "")
            matched.append(row)
    matched.sort(key=lambda r: (date_key(r["date"]), r["no"]))

    wb = Workbook()

    # --- シート1: 抽出結果 ---
    ws = wb.active
    ws.title = "X線装置関連随意契約"
    ws.cell(row=1, column=1, value=f"{hospital} 随意契約公表情報（X線装置関連）").font = Font(bold=True, size=14)

    headers = ["物品などまたは役務の名称", "数量", "随意契約担当課の名称及び所在地",
               "随意契約を締結した日", "随意契約の相手方の氏名及び住所",
               "随意契約に係る契約金額", "URL", "区分（装置カテゴリ）", "製品/保守", "備考"]
    border = style_header(ws, 3, headers)

    r = 4
    for row in matched:
        amount = row["amount"]
        try:
            amount_v = int(re.sub(r"[^0-9]", "", amount)) if re.search(r"\d", amount) else amount
        except ValueError:
            amount_v = amount
        values = [row["name"], row["qty"], row["dept"], row["date"],
                  row["counterparty"], amount_v, row["source_url"],
                  row["category"], row["kind"], row["note"]]
        for ci, v in enumerate(values, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)
            if ci == 6 and isinstance(v, int):
                c.number_format = "#,##0"
        r += 1

    for ci, w in enumerate([40, 6, 34, 14, 34, 14, 48, 30, 10, 40], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A4"

    # --- シート2: 調査サマリー ---
    ws2 = wb.create_sheet("調査サマリー")
    ws2.cell(row=1, column=1, value=f"{hospital} 随意契約公表情報 調査サマリー").font = Font(bold=True, size=14)

    border2 = style_header(ws2, 3, ["装置カテゴリ", "該当件数", "備考"], fill="404040")
    r = 4
    for cat, _ in CATEGORIES:
        n = sum(1 for m in matched if m["category"] == cat)
        note = "" if n else "公表された随意契約に該当なし"
        for ci, v in enumerate([cat, n, note], 1):
            c = ws2.cell(row=r, column=ci, value=v)
            c.border = border2
            c.alignment = Alignment(vertical="top", wrap_text=True)
        r += 1

    r += 1
    ws2.cell(row=r, column=1, value="データソース（随意契約に関する公示 PDF）").font = Font(bold=True)
    r += 1
    style_header(ws2, r, ["対象年度", "URL", "掲載件数"], fill="404040")
    r += 1
    per_src = {}
    for row in rows:
        per_src.setdefault(row["source_url"], []).append(row)
    for url, rr in sorted(per_src.items()):
        dates = sorted((r_["date"] for r_ in rr), key=date_key)
        fys = sorted({fiscal_year(d) for d in dates})
        label = f"{'・'.join(fys)}（{dates[0]}〜{dates[-1]}）"
        for ci, v in enumerate([label, url, len(rr)], 1):
            c = ws2.cell(row=r, column=ci, value=v)
            c.border = border2
            c.alignment = Alignment(vertical="top", wrap_text=True)
        r += 1

    r += 1
    notes = list(EXTRA_NOTES)
    for n_ in notes:
        ws2.cell(row=r, column=1, value=n_)
        r += 1

    for ci, w in enumerate([44, 52, 12], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    wb.save(out_path)
    print(f"saved {out_path}: {len(matched)} matched rows / {len(rows)} total")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2:-1], sys.argv[-1])

"""
レポートを Notion にアップロードするスクリプト
==============================================

output/slack_reaction_report.xlsx の内容を読み取って、
指定した Notion の親ページの下に新しいレポートページを作成します。

やること:
  1. Notion に「Slack リアクションレポート (日付)」ページを新規作成
  2. Excel の3シートの内容を、そのままテーブルとしてページに書き込む
  3. さらに xlsx ファイル本体も Notion にアップロードして添付する
     (アップロードに失敗してもテーブルは残るので安心です)

実行方法:
    python src/upload_to_notion.py

必要な環境変数(.env ファイルに書きます):
    NOTION_TOKEN          : Notion インテグレーションのトークン
    NOTION_PARENT_PAGE_ID : レポートページを作る親ページのID

※ アップロード先の Notion がまだ決まっていない場合は、
   決まってから .env に上記2つを設定すればすぐ使えます。
   親ページに Notion インテグレーションを「接続」するのを忘れずに!
   (ページ右上の「…」→「接続」→ インテグレーション名を選択)
"""

import os
import sys
from datetime import date

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

# 読み込むレポートファイルのパス
REPORT_PATH = os.path.join("output", "slack_reaction_report.xlsx")

# Notion API の基本設定
NOTION_API_BASE = "https://api.notion.com/v1"
NOTION_VERSION = "2022-06-28"  # Notion API のバージョン指定(必須ヘッダー)

# Notion のテーブルに載せる最大行数(ページが長くなりすぎるのを防ぐ)
MAX_TABLE_ROWS = 50


def notion_headers(token):
    """Notion API を呼ぶときに毎回必要なヘッダーをまとめて返す。"""
    return {
        "Authorization": f"Bearer {token}",
        "Notion-Version": NOTION_VERSION,
        "Content-Type": "application/json",
    }


def read_sheet_rows(workbook, sheet_name):
    """Excel のシートから全行を「文字列のリストのリスト」として読み取る。"""
    sheet = workbook[sheet_name]
    rows = []
    for row in sheet.iter_rows(values_only=True):
        # None は空文字にして、すべて文字列に変換する
        rows.append([str(cell) if cell is not None else "" for cell in row])
    return rows


def make_table_block(rows):
    """行データから Notion の「テーブル」ブロックを組み立てる。

    Notion のテーブルは table ブロックの中に table_row ブロックを
    子として入れる構造になっています。
    """
    table_rows = []
    for row in rows[: MAX_TABLE_ROWS + 1]:  # +1 はヘッダー行の分
        table_rows.append({
            "type": "table_row",
            "table_row": {
                # 各セルは rich_text の配列で表現する
                "cells": [
                    [{"type": "text", "text": {"content": cell[:2000]}}]
                    for cell in row
                ],
            },
        })

    return {
        "type": "table",
        "table": {
            "table_width": len(rows[0]),  # 列数
            "has_column_header": True,    # 1行目をヘッダーとして表示
            "children": table_rows,
        },
    }


def make_heading_block(text):
    """見出し(h2)ブロックを組み立てる。"""
    return {
        "type": "heading_2",
        "heading_2": {
            "rich_text": [{"type": "text", "text": {"content": text}}],
        },
    }


def upload_xlsx_file(token):
    """xlsx ファイル本体を Notion の File Upload API でアップロードする。

    成功したらアップロードID(ブロックに添付するときに使う)を返し、
    失敗したら None を返す。
    """
    try:
        # ステップ1: アップロード枠を作成する
        response = requests.post(
            f"{NOTION_API_BASE}/file_uploads",
            headers=notion_headers(token),
            json={
                "filename": os.path.basename(REPORT_PATH),
                "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            },
            timeout=30,
        )
        response.raise_for_status()
        upload = response.json()

        # ステップ2: 実際のファイルを送信する
        with open(REPORT_PATH, "rb") as f:
            send_response = requests.post(
                f"{NOTION_API_BASE}/file_uploads/{upload['id']}/send",
                headers={
                    "Authorization": f"Bearer {token}",
                    "Notion-Version": NOTION_VERSION,
                },
                files={"file": (os.path.basename(REPORT_PATH), f)},
                timeout=60,
            )
        send_response.raise_for_status()
        return upload["id"]

    except requests.RequestException as e:
        # ファイル添付はおまけ機能なので、失敗しても処理は続行する
        print(f"  注意: xlsx ファイルの添付に失敗しました({e})")
        print("  テーブルだけでページを作成します。")
        return None


def main():
    # .env ファイルから環境変数を読み込む
    load_dotenv()

    token = os.environ.get("NOTION_TOKEN")
    parent_page_id = os.environ.get("NOTION_PARENT_PAGE_ID")

    # 必要な環境変数が揃っているかチェックする
    if not token or not parent_page_id:
        print("エラー: NOTION_TOKEN と NOTION_PARENT_PAGE_ID を .env に設定してください。")
        print("(アップロード先の Notion が決まったら設定すれば使えます)")
        sys.exit(1)

    # レポートファイルが存在するかチェックする
    if not os.path.exists(REPORT_PATH):
        print(f"エラー: {REPORT_PATH} が見つかりません。")
        print("先に python src/slack_reaction_report.py を実行してください。")
        sys.exit(1)

    # ------------------------------------------------------------
    # Excel からデータを読み込む
    # ------------------------------------------------------------
    workbook = load_workbook(REPORT_PATH, read_only=True)

    # シート名と Notion 上の見出しの対応表
    sheets = [
        ("most_reactive_users", "リアクションした回数ランキング"),
        ("most_reacted_users", "リアクションされた回数ランキング"),
        ("emoji_ranking", "絵文字ランキング"),
    ]

    # ページに入れるブロックを順番に組み立てる
    blocks = []
    for sheet_name, heading in sheets:
        rows = read_sheet_rows(workbook, sheet_name)
        blocks.append(make_heading_block(heading))
        if len(rows) > 1:  # ヘッダー以外にデータ行がある場合だけテーブルを作る
            blocks.append(make_table_block(rows))
        else:
            blocks.append({
                "type": "paragraph",
                "paragraph": {
                    "rich_text": [{"type": "text", "text": {"content": "データがありません"}}],
                },
            })

    # ------------------------------------------------------------
    # xlsx ファイル本体もアップロードして添付ブロックを追加する
    # ------------------------------------------------------------
    print("xlsx ファイルを Notion にアップロードしています...")
    file_upload_id = upload_xlsx_file(token)
    if file_upload_id:
        blocks.append(make_heading_block("Excel ファイル"))
        blocks.append({
            "type": "file",
            "file": {
                "type": "file_upload",
                "file_upload": {"id": file_upload_id},
            },
        })

    # ------------------------------------------------------------
    # 親ページの下に新しいページを作成する
    # ------------------------------------------------------------
    today = date.today().strftime("%Y-%m-%d")
    print("Notion にレポートページを作成しています...")

    try:
        response = requests.post(
            f"{NOTION_API_BASE}/pages",
            headers=notion_headers(token),
            json={
                "parent": {"page_id": parent_page_id},
                "properties": {
                    "title": {
                        "title": [{
                            "type": "text",
                            "text": {"content": f"Slack リアクションレポート ({today})"},
                        }],
                    },
                },
                "children": blocks,
            },
            timeout=60,
        )
        response.raise_for_status()
        page_url = response.json().get("url", "")
        print(f"完了! Notion ページを作成しました: {page_url}")

    except requests.HTTPError as e:
        status = e.response.status_code
        print(f"エラー: Notion ページの作成に失敗しました (HTTP {status})")
        if status == 401:
            print("→ NOTION_TOKEN が正しいか確認してください。")
        elif status == 404:
            print("→ NOTION_PARENT_PAGE_ID が正しいか、")
            print("  親ページにインテグレーションが「接続」されているか確認してください。")
        else:
            print(f"→ 詳細: {e.response.text[:500]}")
        sys.exit(1)


if __name__ == "__main__":
    main()

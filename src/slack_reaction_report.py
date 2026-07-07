"""
Slack リアクション分析ツール
============================

指定した Slack チャンネルの過去30日分のメッセージを取得して、
以下の3つの集計を Excel (output/slack_reaction_report.xlsx) に出力します。

1. most_reactive_users : リアクションを「した」回数が多いユーザーランキング
2. most_reacted_users  : リアクションを「された」回数が多いユーザーランキング
3. emoji_ranking       : 使われた絵文字のランキング

実行方法:
    python src/slack_reaction_report.py

必要な環境変数(.env ファイルに書きます):
    SLACK_BOT_TOKEN  : Slack Bot のトークン (xoxb- で始まる)
    SLACK_CHANNEL_ID : 対象チャンネルID(省略時は CBHRRSZAP)

必要な Bot スコープ:
    channels:history, channels:read, users:read
    (プライベートチャンネルの場合は groups:history, groups:read も)
"""

import os
import sys
import time
from collections import Counter
from datetime import datetime, timedelta, timezone

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font
from slack_sdk import WebClient
from slack_sdk.errors import SlackApiError

# ------------------------------------------------------------
# 設定値
# ------------------------------------------------------------

# 集計対象の日数(過去30日)
DAYS_TO_ANALYZE = 30

# デフォルトの対象チャンネルID(環境変数で上書き可能)
DEFAULT_CHANNEL_ID = "CBHRRSZAP"

# 出力先の Excel ファイルパス
OUTPUT_PATH = os.path.join("output", "slack_reaction_report.xlsx")

# API を連続で呼ぶときの待ち時間(秒)。rate limit 対策です。
API_CALL_INTERVAL = 1.1

# rate limit (429エラー) が返ってきたときの最大リトライ回数
MAX_RETRIES = 5


# ------------------------------------------------------------
# Slack API を安全に呼ぶためのヘルパー関数
# ------------------------------------------------------------

def call_slack_api(api_method, **kwargs):
    """Slack API を呼び出す。rate limit (429) の場合は待ってからリトライする。

    api_method : client.conversations_history のような関数
    kwargs     : API に渡すパラメータ
    """
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            return api_method(**kwargs)
        except SlackApiError as e:
            # 429 = rate limit(呼びすぎ)。Retry-After ヘッダの秒数だけ待つ
            if e.response.status_code == 429:
                wait_seconds = int(e.response.headers.get("Retry-After", 10))
                print(f"  rate limit に達しました。{wait_seconds}秒待ちます... "
                      f"(リトライ {attempt}/{MAX_RETRIES})")
                time.sleep(wait_seconds)
            else:
                # 429 以外のエラーはリトライしても直らないのでそのまま上に投げる
                raise
    raise RuntimeError(f"{MAX_RETRIES}回リトライしましたが rate limit が解消しませんでした")


# ------------------------------------------------------------
# ユーザー情報の取得 (users.list)
# ------------------------------------------------------------

def fetch_users(client):
    """ワークスペースの全ユーザーを取得して、
    「ユーザーID → 表示名」の辞書と「botのユーザーIDの集合」を返す。
    """
    user_names = {}   # 例: {"U12345": "tanaka"}
    bot_user_ids = set()  # bot のユーザーID一覧

    cursor = None  # ページネーション用のカーソル
    while True:
        response = call_slack_api(
            client.users_list,
            limit=200,
            cursor=cursor,
        )
        for user in response["members"]:
            user_id = user["id"]

            # 表示名の優先順位: display_name → real_name → name(ログイン名)
            profile = user.get("profile", {})
            name = (
                profile.get("display_name")
                or profile.get("real_name")
                or user.get("name")
                or user_id
            )
            user_names[user_id] = name

            # bot か Slackbot なら bot 集合に入れる(あとで除外するため)
            if user.get("is_bot") or user_id == "USLACKBOT":
                bot_user_ids.add(user_id)

        # next_cursor が空なら最後のページなので終了
        cursor = response.get("response_metadata", {}).get("next_cursor")
        if not cursor:
            break
        time.sleep(API_CALL_INTERVAL)

    return user_names, bot_user_ids


# ------------------------------------------------------------
# メッセージの取得 (conversations.history + conversations.replies)
# ------------------------------------------------------------

def fetch_channel_messages(client, channel_id, oldest_ts):
    """チャンネルのメッセージを過去30日分すべて取得する(ページネーション対応)。"""
    messages = []
    cursor = None
    page = 0

    while True:
        page += 1
        print(f"  メッセージ取得中... (ページ {page})")
        response = call_slack_api(
            client.conversations_history,
            channel=channel_id,
            oldest=oldest_ts,   # この時刻より新しいメッセージだけ取得
            limit=200,          # 1回で取る最大件数
            cursor=cursor,
        )
        messages.extend(response["messages"])

        cursor = response.get("response_metadata", {}).get("next_cursor")
        if not cursor:
            break
        time.sleep(API_CALL_INTERVAL)

    return messages


def fetch_thread_replies(client, channel_id, thread_ts, oldest_ts):
    """スレッドの返信メッセージをすべて取得する(ページネーション対応)。

    conversations.replies は「親メッセージ + 返信」を返すので、
    親メッセージ(ts == thread_ts)は除外して返信だけを返します。
    """
    replies = []
    cursor = None

    while True:
        response = call_slack_api(
            client.conversations_replies,
            channel=channel_id,
            ts=thread_ts,
            oldest=oldest_ts,
            limit=200,
            cursor=cursor,
        )
        for message in response["messages"]:
            # 親メッセージは conversations.history で取得済みなのでスキップ
            if message.get("ts") != thread_ts:
                replies.append(message)

        cursor = response.get("response_metadata", {}).get("next_cursor")
        if not cursor:
            break
        time.sleep(API_CALL_INTERVAL)

    return replies


def is_bot_message(message, bot_user_ids):
    """このメッセージが bot の投稿かどうかを判定する。"""
    # bot_id があれば bot の投稿
    if message.get("bot_id"):
        return True
    # subtype が bot_message の場合も bot の投稿
    if message.get("subtype") == "bot_message":
        return True
    # 投稿者のユーザーIDが bot 一覧に含まれる場合
    if message.get("user") in bot_user_ids:
        return True
    return False


# ------------------------------------------------------------
# リアクションの集計
# ------------------------------------------------------------

def aggregate_reactions(messages, bot_user_ids):
    """メッセージ一覧からリアクションを集計する。

    戻り値(4つの Counter):
        reactions_made     : ユーザーID → リアクションした回数
        reactions_received : ユーザーID → リアクションされた回数
        emoji_by_maker     : (ユーザーID, 絵文字) → 回数(よく使う絵文字の算出用)
        emoji_by_receiver  : (ユーザーID, 絵文字) → 回数(よくもらう絵文字の算出用)
        emoji_total        : 絵文字 → 全体での使用回数
    """
    reactions_made = Counter()
    reactions_received = Counter()
    emoji_by_maker = Counter()
    emoji_by_receiver = Counter()
    emoji_total = Counter()

    for message in messages:
        # bot の投稿は集計から除外する(要件)
        if is_bot_message(message, bot_user_ids):
            continue

        author_id = message.get("user")  # メッセージの投稿者
        reactions = message.get("reactions", [])  # リアクションが無ければ空リスト

        for reaction in reactions:
            emoji_name = reaction["name"]       # 例: "thumbsup"
            reacted_users = reaction.get("users", [])  # リアクションした人のID一覧

            for reactor_id in reacted_users:
                # bot がしたリアクションは除外
                if reactor_id in bot_user_ids:
                    continue

                # 「リアクションした」側の集計
                reactions_made[reactor_id] += 1
                emoji_by_maker[(reactor_id, emoji_name)] += 1

                # 「リアクションされた」側の集計(投稿者ごと)
                if author_id:
                    reactions_received[author_id] += 1
                    emoji_by_receiver[(author_id, emoji_name)] += 1

                # 絵文字全体のランキング用
                emoji_total[emoji_name] += 1

    return (reactions_made, reactions_received,
            emoji_by_maker, emoji_by_receiver, emoji_total)


def find_top_emoji(emoji_counter, user_id):
    """指定ユーザーが一番よく使った(またはもらった)絵文字を返す。"""
    best_emoji = ""
    best_count = 0
    for (uid, emoji), count in emoji_counter.items():
        if uid == user_id and count > best_count:
            best_emoji = emoji
            best_count = count
    return best_emoji


# ------------------------------------------------------------
# Excel 出力
# ------------------------------------------------------------

def write_excel(user_names, reactions_made, reactions_received,
                emoji_by_maker, emoji_by_receiver, emoji_total):
    """集計結果を3枚のシートに分けて Excel ファイルに書き出す。"""
    workbook = Workbook()
    bold = Font(bold=True)  # ヘッダー行を太字にするためのスタイル

    # --- シート1: リアクションした回数ランキング ---
    sheet1 = workbook.active
    sheet1.title = "most_reactive_users"
    sheet1.append(["rank", "user_name", "reactions_made", "favorite_emoji"])
    for cell in sheet1[1]:
        cell.font = bold

    # most_common() で回数が多い順に並べる
    for rank, (user_id, count) in enumerate(reactions_made.most_common(), start=1):
        sheet1.append([
            rank,
            user_names.get(user_id, user_id),  # 名前が引けない場合はIDのまま表示
            count,
            find_top_emoji(emoji_by_maker, user_id),
        ])

    # --- シート2: リアクションされた回数ランキング ---
    sheet2 = workbook.create_sheet("most_reacted_users")
    sheet2.append(["rank", "user_name", "reactions_received", "top_received_emoji"])
    for cell in sheet2[1]:
        cell.font = bold

    for rank, (user_id, count) in enumerate(reactions_received.most_common(), start=1):
        sheet2.append([
            rank,
            user_names.get(user_id, user_id),
            count,
            find_top_emoji(emoji_by_receiver, user_id),
        ])

    # --- シート3: 絵文字ランキング ---
    sheet3 = workbook.create_sheet("emoji_ranking")
    sheet3.append(["emoji", "count"])
    for cell in sheet3[1]:
        cell.font = bold

    for emoji, count in emoji_total.most_common():
        sheet3.append([emoji, count])

    # 列幅を少し広げて見やすくする
    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            letter = column_cells[0].column_letter
            sheet.column_dimensions[letter].width = 22

    # output フォルダが無ければ作る
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    workbook.save(OUTPUT_PATH)


# ------------------------------------------------------------
# メイン処理
# ------------------------------------------------------------

def main():
    # .env ファイルから環境変数を読み込む
    load_dotenv()

    token = os.environ.get("SLACK_BOT_TOKEN")
    if not token:
        print("エラー: 環境変数 SLACK_BOT_TOKEN が設定されていません。")
        print(".env ファイルを作成して SLACK_BOT_TOKEN を設定してください。")
        sys.exit(1)

    channel_id = os.environ.get("SLACK_CHANNEL_ID", DEFAULT_CHANNEL_ID)
    client = WebClient(token=token)

    # 「30日前」の UNIXタイムスタンプを計算する
    oldest_datetime = datetime.now(timezone.utc) - timedelta(days=DAYS_TO_ANALYZE)
    oldest_ts = str(oldest_datetime.timestamp())

    print(f"対象チャンネル : {channel_id}")
    print(f"集計期間       : {oldest_datetime.strftime('%Y-%m-%d')} 〜 今日(過去{DAYS_TO_ANALYZE}日間)")

    try:
        # ステップ1: ユーザー一覧を取得(名前の解決と bot 判定に使う)
        print("\n[1/4] ユーザー一覧を取得しています...")
        user_names, bot_user_ids = fetch_users(client)
        print(f"  {len(user_names)}人のユーザーを取得しました(うち bot: {len(bot_user_ids)})")

        # ステップ2: チャンネルのメッセージを取得
        print("\n[2/4] チャンネルのメッセージを取得しています...")
        messages = fetch_channel_messages(client, channel_id, oldest_ts)
        print(f"  {len(messages)}件のメッセージを取得しました")

        # ステップ3: スレッドの返信も取得して追加する(要件: thread返信も含める)
        print("\n[3/4] スレッドの返信を取得しています...")
        thread_replies = []
        for message in messages:
            # thread_ts == ts のメッセージが「スレッドの親」
            # reply_count があれば返信が存在する
            is_thread_parent = (
                message.get("thread_ts") == message.get("ts")
                and message.get("reply_count", 0) > 0
            )
            if is_thread_parent:
                replies = fetch_thread_replies(
                    client, channel_id, message["ts"], oldest_ts
                )
                thread_replies.extend(replies)
                time.sleep(API_CALL_INTERVAL)
        print(f"  {len(thread_replies)}件のスレッド返信を取得しました")

        all_messages = messages + thread_replies

        # ステップ4: 集計して Excel に出力
        print("\n[4/4] 集計して Excel に出力しています...")
        (reactions_made, reactions_received,
         emoji_by_maker, emoji_by_receiver, emoji_total) = aggregate_reactions(
            all_messages, bot_user_ids
        )
        write_excel(user_names, reactions_made, reactions_received,
                    emoji_by_maker, emoji_by_receiver, emoji_total)

        print(f"\n完了! レポートを出力しました: {OUTPUT_PATH}")
        print(f"  リアクションしたユーザー数 : {len(reactions_made)}")
        print(f"  リアクションされたユーザー数: {len(reactions_received)}")
        print(f"  絵文字の種類               : {len(emoji_total)}")

    except SlackApiError as e:
        # Slack API のエラー内容をわかりやすく表示する
        error_code = e.response.get("error", "unknown")
        print(f"\nSlack API エラーが発生しました: {error_code}")
        if error_code == "invalid_auth":
            print("→ SLACK_BOT_TOKEN が正しいか確認してください。")
        elif error_code == "channel_not_found":
            print(f"→ チャンネル {channel_id} が見つかりません。IDを確認してください。")
        elif error_code == "not_in_channel":
            print(f"→ Bot をチャンネルに招待してください(/invite @ボット名)。")
        elif error_code == "missing_scope":
            print("→ Bot に必要なスコープ(channels:history, users:read など)を追加してください。")
        sys.exit(1)


if __name__ == "__main__":
    main()

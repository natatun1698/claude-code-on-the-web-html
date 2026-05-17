import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 全病院の放射線機器データ
data = [
    # 病院名, 機器種別, メーカー名, 型名, 導入年月
    # グループ1
    ("副都心病院", "一般撮影", "", "", ""),
    ("副都心病院", "透視撮影台（X線TV）", "", "", ""),
    ("としま昭和病院", "一般撮影", "", "", ""),
    ("としま昭和病院", "透視撮影台（X線TV）", "", "", ""),
    ("要町病院", "一般撮影", "", "", ""),
    ("要町病院", "透視撮影台（X線TV）", "", "", ""),
    ("高田馬場病院", "一般撮影", "", "", ""),
    ("高田馬場病院", "透視撮影台（X線TV）", "", "", ""),
    ("大同病院", "一般撮影", "", "", ""),
    ("大同病院", "透視撮影台（X線TV）", "", "", ""),
    ("大同病院", "血管撮影（アンギオ）", "Philips", "Xper CT / Xper Guide搭載 大視野3D", ""),
    ("関野病院", "一般撮影", "", "", ""),
    ("関野病院", "透視撮影台（X線TV）", "", "", ""),
    ("長汐病院", "一般撮影", "", "", ""),
    ("長汐病院", "透視撮影台（X線TV）", "", "", ""),
    ("池袋西口病院", "一般撮影", "", "", ""),
    ("池袋西口病院", "透視撮影台（X線TV）", "", "", ""),
    ("池袋病院", "一般撮影", "", "", ""),
    ("池袋病院", "透視撮影台（X線TV）", "", "", ""),
    ("豊島中央病院", "一般撮影", "", "", ""),
    ("豊島中央病院", "透視撮影台（X線TV）", "", "", ""),
    # グループ2
    ("一心病院", "一般撮影", "", "", ""),
    ("一心病院", "透視撮影台（X線TV）", "富士フイルム", "CUREVISTA（要確認）", ""),
    ("浅草寺病院", "一般撮影", "", "", ""),
    ("浅草寺病院", "透視撮影台（X線TV）", "", "", ""),
    ("同善病院", "一般撮影", "", "", ""),
    ("同善病院", "透視撮影台（X線TV）", "", "", ""),
    ("神経科 土田病院", "一般撮影", "", "", ""),
    ("慈誠会・光が丘病院", "一般撮影", "", "", ""),
    ("慈誠会・光が丘病院", "透視撮影台（X線TV）", "", "", ""),
    ("人間ドック会館クリニック", "一般撮影", "", "", ""),
    ("成増病院", "一般撮影", "", "", ""),
    ("成増病院", "透視撮影台（X線TV）", "", "", ""),
    ("慈誠会記念病院", "一般撮影", "", "", ""),
    ("慈誠会記念病院", "透視撮影台（X線TV）", "", "", ""),
    ("小林病院", "一般撮影", "", "", ""),
    ("小林病院", "透視撮影台（X線TV）", "", "", ""),
    ("安田病院", "一般撮影", "", "", ""),
    ("安田病院", "透視撮影台（X線TV）", "", "", ""),
    # グループ3
    ("成増厚生病院", "一般撮影", "", "", ""),
    ("成増厚生病院", "透視撮影台（X線TV）", "", "", ""),
    ("東武練馬中央病院", "一般撮影", "", "", ""),
    ("東武練馬中央病院", "透視撮影台（X線TV）", "", "", ""),
    ("慈誠会徳丸リハビリテーション病院", "一般撮影", "", "", ""),
    ("板橋区医師会病院", "一般撮影", "", "", ""),
    ("板橋区医師会病院", "透視撮影台（X線TV）", "キヤノン（要確認）", "", ""),
    ("上板橋病院", "一般撮影", "", "", ""),
    ("上板橋病院", "透視撮影台（X線TV）", "", "", ""),
    ("飯沼病院", "一般撮影", "", "", ""),
    # グループ4
    ("小豆沢病院", "一般撮影", "", "", ""),
    ("小豆沢病院", "透視撮影台（X線TV）", "", "", ""),
    ("金子病院", "一般撮影", "", "", ""),
    ("金子病院", "透視撮影台（X線TV）", "", "", ""),
    ("常盤台外科病院", "一般撮影", "", "", ""),
    ("常盤台外科病院", "透視撮影台（X線TV）", "", "", ""),
    ("慈誠会若木原病院", "一般撮影", "", "", ""),
    ("慈誠会若木原病院", "透視撮影台（X線TV）", "", "", ""),
    ("常盤台病院", "一般撮影", "", "", ""),
    ("常盤台病院", "透視撮影台（X線TV）", "", "", ""),
    ("慈誠会前野病院", "一般撮影", "", "", ""),
    ("慈誠会前野病院", "透視撮影台（X線TV）", "", "", ""),
    ("板橋宮本病院", "一般撮影", "", "", ""),
    ("板橋宮本病院", "透視撮影台（X線TV）", "", "", ""),
    ("誠志会病院", "一般撮影", "", "", ""),
    ("誠志会病院", "透視撮影台（X線TV）", "", "", ""),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "放射線機器一覧"

# スタイル定義
header_font = Font(name="MS Gothic", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
note_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
body_font = Font(name="MS Gothic", size=10)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
body_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin = Side(style="thin", color="BFBFBF")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ヘッダー行
headers = ["病院名", "機器種別", "メーカー名", "型名", "導入年月", "備考"]
ws.append(["放射線機器（レントゲン）導入状況調査", "", "", "", "", ""])
ws.merge_cells("A1:F1")
ws["A1"].font = Font(name="MS Gothic", bold=True, size=14, color="FFFFFF")
ws["A1"].fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

ws.append(["対象機器：一般撮影・透視撮影台（X線TV）・血管撮影（アンギオ）", "", "", "", "", ""])
ws.merge_cells("A2:F2")
ws["A2"].font = Font(name="MS Gothic", size=10, color="FFFFFF")
ws["A2"].fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 18

ws.append(headers)
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_num)
    cell.font = header_font
    cell.fill = subheader_fill
    cell.alignment = header_align
    cell.border = thin_border
ws.row_dimensions[3].height = 22

# データ行
prev_hospital = None
row_num = 4
hospital_color_toggle = True

for row_data in data:
    hospital, equip_type, maker, model, date = row_data

    if hospital != prev_hospital:
        hospital_color_toggle = not hospital_color_toggle
        prev_hospital = hospital

    note = ""
    if maker and "要確認" in maker:
        note = "ウェブ検索による断片的情報。要確認。"
    elif model and "要確認" in model:
        note = "ウェブ検索による断片的情報。要確認。"

    row_values = [hospital, equip_type, maker, model, date, note]
    ws.append(row_values)

    for col_num, val in enumerate(row_values, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.font = body_font
        cell.border = thin_border
        cell.alignment = body_align

        if note and col_num in [3, 4, 6]:
            cell.fill = note_fill
        elif hospital_color_toggle:
            cell.fill = alt_fill

    ws.row_dimensions[row_num].height = 18
    row_num += 1

# 列幅設定
col_widths = [22, 20, 20, 30, 12, 35]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# ウィンドウ枠の固定
ws.freeze_panes = "A4"

# 注記シート
ws2 = wb.create_sheet("調査メモ")
notes = [
    ["調査日", "2026年5月17日"],
    ["調査方法", "各病院公式ウェブサイトの設備案内ページを参照"],
    [""],
    ["調査結果サマリー"],
    ["対象病院数", "34病院"],
    ["情報取得状況", "大多数の病院はウェブ上に機器の詳細仕様（メーカー・型番・導入年月）を非公開"],
    [""],
    ["断片的情報が確認された病院"],
    ["大同病院", "血管撮影装置にPhilips Xper搭載との記述あり（型番・導入年月は要確認）"],
    ["一心病院", "透視撮影台に富士フイルム CUREVISTA搭載との記述あり（要確認）"],
    ["板橋区医師会病院", "透視撮影台3台がキヤノン製との記述あり（型番・導入年月は要確認）"],
    [""],
    ["推奨する次のステップ"],
    ["1", "各病院の放射線科・医事課に直接電話で問い合わせる"],
    ["2", "厚生労働省「医療情報ネット（ナビイ）」で高額医療機器情報を確認する"],
    ["3", "病院への訪問・見学を通じて設備情報を収集する"],
]

for i, row in enumerate(notes, 1):
    ws2.append(row)
    if row and row[0] in ["調査結果サマリー", "断片的情報が確認された病院", "推奨する次のステップ"]:
        ws2.cell(row=i, column=1).font = Font(name="MS Gothic", bold=True, size=11)
    else:
        for j in range(1, 3):
            ws2.cell(row=i, column=j).font = Font(name="MS Gothic", size=10)

ws2.column_dimensions["A"].width = 25
ws2.column_dimensions["B"].width = 60

output_path = "/home/user/claude-code-on-the-web-html/病院放射線機器一覧.xlsx"
wb.save(output_path)
print(f"Excel file saved: {output_path}")
